import os, time, math, openpyxl, datetime
from openpyxl.styles import PatternFill, Font, Alignment
from dependencies.market_search_modules import ownerquery_details, start_ownerquery_details
from dependencies.gene_class import GetGenes256
maxsleep = 2
axie_list = []
f = open("txt_files/ronin.txt")
ronin_list = [x.strip() for x in f.readlines() if x != ""]
ronin_list = [x.replace("ronin:", "0x") for x in ronin_list]
axie_list = []
wb = openpyxl.load_workbook("templates/ronin_axie_details.xlsx")
ws = wb['Sheet1']
hyperlinked = Font(underline = "single", color="0C10F5")
striked = Font(strike = True, color = "0C10F5")

colors = {"beast": "FFB812", "aquatic": "00B8CE", "bird": "FF8BBD", "plant": "6CC000", "bug": "ff5341", "reptile": "B740CF", "dusk": "129092", "mech": "C6BDD4", "daw": "BECEFF"}


for ronin in ronin_list:
	while True:
		try:
			time.sleep(maxsleep)
			data = start_ownerquery_details(ronin, index = 0)
			break
		except ConnectionError:
			time.sleep(maxsleep)
			print("Server Connection Error. Retrying")

	try:
		total_axies = data['data']['axies']['total']
		print(ronin + ' Total '+ str(total_axies))
		if total_axies <= 0:
			pass
		else:
			loops = math.ceil(total_axies / 100)
			for i in range(loops + 1):
				while True:
					try:
						time.sleep(1)
						data = start_ownerquery_details(ronin, index = i * 100)
						break
					except ConnectionError:
						time.sleep(1)
						print("Server Connection Error. Retrying")
				for axie in data['data']['axies']['results']:
					axie_list.append(axie)
	except KeyError:
		pass

for axie in axie_list:
	# print("axie ID: " + axie['id'] + " Owner: " + axie['owner'])
	actual_birthdate = axie['birthDate'] + 432000
	string_birthdate = datetime.datetime.fromtimestamp(actual_birthdate).strftime("%m/%d/%Y %I:%M %p")
	max_row = ws.max_row + 1
	if axie['stage'] != 1:
		genes = GetGenes256(axie['genes'])
		ws.cell(row = max_row, column = 1).value = axie['id']
		if axie['auction'] is None:
			ws.cell(row = max_row, column = 1).font = hyperlinked
		else:
			ws.cell(row = max_row, column = 1).font = striked
		ws.cell(row = max_row, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['id']}/"
		ws.cell(row = max_row, column = 2).value = axie['class']
		ws.cell(row = max_row, column = 2).font = Font(color = colors[axie['class'].lower()])
		ws.cell(row = max_row, column = 3).value = axie['owner']
		ws.cell(row = max_row, column = 4).value = genes.purity
		ws.cell(row = max_row, column = 5).value = axie['breedCount']
		ws.cell(row = max_row, column = 6).value = string_birthdate
		ws.cell(row = max_row, column = 6).number_format = "dd/mm/yy hh:mm AM/PM"
		ws.cell(row = max_row, column = 6).alignment = Alignment(horizontal = "right")
		ws.cell(row = max_row, column = 7).value = axie['sireId']
		ws.cell(row = max_row, column = 7).font = hyperlinked
		ws.cell(row = max_row, column = 7).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['sireId']}/"
		ws.cell(row = max_row, column = 8).value = axie['matronId']
		ws.cell(row = max_row, column = 8).font = hyperlinked
		ws.cell(row = max_row, column = 8).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['matronId']}/"

		#d
		ws.cell(row = max_row, column = 9).value = genes.eyes.dName.title()
		ws.cell(row = max_row, column = 9).font = Font(color = colors[genes.eyes.dClass.lower()])
		ws.cell(row = max_row, column = 10).value = genes.ears.dName.title()
		ws.cell(row = max_row, column = 10).font = Font(color = colors[genes.ears.dClass.lower()])
		ws.cell(row = max_row, column = 11).value = genes.horn.dName.title()
		ws.cell(row = max_row, column = 11).font = Font(color = colors[genes.horn.dClass.lower()])
		ws.cell(row = max_row, column = 12).value = genes.mouth.dName.title()
		ws.cell(row = max_row, column = 12).font = Font(color = colors[genes.mouth.dClass.lower()])
		ws.cell(row = max_row, column = 13).value = genes.back.dName.title()
		ws.cell(row = max_row, column = 13).font = Font(color = colors[genes.back.dClass.lower()])
		ws.cell(row = max_row, column = 14).value = genes.tail.dName.title()
		ws.cell(row = max_row, column = 14).font = Font(color = colors[genes.tail.dClass.lower()])

		#r1
		ws.cell(row = max_row, column = 15).value = genes.eyes.r1Name.title()
		ws.cell(row = max_row, column = 15).font = Font(color = colors[genes.eyes.r1Class.lower()])
		ws.cell(row = max_row, column = 16).value = genes.ears.r1Name.title()
		ws.cell(row = max_row, column = 16).font = Font(color = colors[genes.ears.r1Class.lower()])
		ws.cell(row = max_row, column = 17).value = genes.horn.r1Name.title()
		ws.cell(row = max_row, column = 17).font = Font(color = colors[genes.horn.r1Class.lower()])
		ws.cell(row = max_row, column = 18).value = genes.mouth.r1Name.title()
		ws.cell(row = max_row, column = 18).font = Font(color = colors[genes.mouth.r1Class.lower()])
		ws.cell(row = max_row, column = 19).value = genes.back.r1Name.title()
		ws.cell(row = max_row, column = 19).font = Font(color = colors[genes.back.r1Class.lower()])
		ws.cell(row = max_row, column = 20).value = genes.tail.r1Name.title()
		ws.cell(row = max_row, column = 20).font = Font(color = colors[genes.tail.r1Class.lower()])

		#r2
		ws.cell(row = max_row, column = 21).value = genes.eyes.r2Name.title()
		ws.cell(row = max_row, column = 21).font = Font(color = colors[genes.eyes.r2Class.lower()])
		ws.cell(row = max_row, column = 22).value = genes.ears.r2Name.title()
		ws.cell(row = max_row, column = 22).font = Font(color = colors[genes.ears.r2Class.lower()])
		ws.cell(row = max_row, column = 23).value = genes.horn.r2Name.title()
		ws.cell(row = max_row, column = 23).font = Font(color = colors[genes.horn.r2Class.lower()])
		ws.cell(row = max_row, column = 24).value = genes.mouth.r2Name.title()
		ws.cell(row = max_row, column = 24).font = Font(color = colors[genes.mouth.r2Class.lower()])
		ws.cell(row = max_row, column = 25).value = genes.back.r2Name.title()
		ws.cell(row = max_row, column = 25).font = Font(color = colors[genes.back.r2Class.lower()])
		ws.cell(row = max_row, column = 26).value = genes.tail.r2Name.title()
		ws.cell(row = max_row, column = 26).font = Font(color = colors[genes.tail.r2Class.lower()])

		#stats
		ws.cell(row = max_row, column = 27).value = axie['stats']['hp']
		ws.cell(row = max_row, column = 28).value = axie['stats']['speed']
		ws.cell(row = max_row, column = 29).value = axie['stats']['skill']
		ws.cell(row = max_row, column = 30).value = axie['stats']['morale']

	else:
		ws.cell(row = max_row, column = 1).value = axie['id']
		ws.cell(row = max_row, column = 1).font = hyperlinked
		ws.cell(row = max_row, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['id']}/"
		ws.cell(row = max_row, column = 2).value = "Egg"
		ws.cell(row = max_row, column = 3).value = axie['owner']
		ws.cell(row = max_row, column = 5).value = 0
		ws.cell(row = max_row, column = 6).value = string_birthdate
		ws.cell(row = max_row, column = 6).number_format = "dd/mm/yy hh:mm AM/PM"
		ws.cell(row = max_row, column = 6).alignment = Alignment(horizontal = "right")
		ws.cell(row = max_row, column = 7).value = axie['sireId']
		ws.cell(row = max_row, column = 7).font = hyperlinked
		ws.cell(row = max_row, column = 7).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['sireId']}/"
		ws.cell(row = max_row, column = 8).value = axie['matronId']
		ws.cell(row = max_row, column = 8).font = hyperlinked
		ws.cell(row = max_row, column = 8).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['matronId']}/"



for i, row in enumerate(ws.iter_rows(min_row = 3)):
	if i % 2 == 0:
		for cell in row:
			cell.fill = PatternFill(fill_type="solid", start_color = "c8f4fa", end_color = "c8f4fa")

ws.auto_filter.ref = "A2:AD2"

tday = datetime.datetime.today().strftime("%m_%d_%Y")
dirlist = os.listdir("outputs")
pile = [x for x in dirlist if f"ronin_axie_details_{tday}" in x]
pile = [x.replace(f"ronin_axie_details_{tday}_", "").replace(".xlsx", "").replace("~$", "") for x in pile]
counts = [int(x) for x in pile]
if len(counts) <= 0:
	d = 1
else:
	d = max(counts) + 1


wb.save(f"outputs/ronin_axie_details_{tday}_{d}.xlsx")
print("finished")


		