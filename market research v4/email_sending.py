import smtplib, ssl, json, time, datetime, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from email.message import EmailMessage
from dependencies.market_search_modules import start_marketquery_2
from dependencies.mymodules import convert_to_eth
from dependencies.gene_class import GetGenes256
import threading

colors = {"beast": "FFB812", "aquatic": "00B8CE", "bird": "FF8BBD", "plant": "6CC000", "bug": "ff5341", "reptile": "B740CF", "dusk": "129092", "mech": "C6BDD4", "daw": "BECEFF"}
hyperlinked = Font(underline = "single", color="0C10F5")
sleep_time = 4
parts_file = open("json_files/parts.json")
parts_json = json.load(parts_file)
parts_file.close()
f = open("json_files/search.json")
f_json = json.load(f)
delay = f_json['delay']
receiver = f_json['email']
filters = f_json['filters']
search_filters = {}

def run_program():
	for _filter in filters:
		_name = _filter['filter_name']
		search_filters[_name] = {}
		d_parts = _filter['d.parts']
		search_filters[_name]['d'] = d_parts
		try:
			breedCount = _filter['breedCount']
			breedCount = [0, breedCount]
		except KeyError:
			breedCount = [0, 6]

		search_filters[_name]['breedCount'] = breedCount

		try:
			r1_parts = _filter['r1.parts']
			search_filters[_name]['r1'] = r1_parts
		except KeyError:
			search_filters[_name]['r1'] = []
		try:
			r2_parts = _filter['r2.parts']
			search_filters[_name]['r2'] = r2_parts
		except KeyError:
			search_filters[_name]['r2'] = []

		price = _filter['price_eth']
		try:
			axie_class = [_filter['class'].title()]
		except KeyError:
			axie_class = []

		search_filters[_name]['class'] = axie_class

		try:
			hp = _filter['hp']
			hp = [hp, 61]
		except KeyError:
			hp = [0, 61]

		search_filters[_name]['hp'] = hp

		try:
			speed = _filter['speed']
			speed = [speed, 61]
		except KeyError:
			speed = [0, 61]

		search_filters[_name]['speed'] = speed

		try:
			skill = _filter['skill']
			skill = [skill, 61]
		except KeyError:
			skill = [0, 61]
		
		search_filters[_name]['skill'] = skill

		try:
			morale = _filter['morale']
			morale = [morale, 61]
		except KeyError:
			morale = [0, 61]
		search_filters[_name]['morale'] = morale

		search_filters[_name]['price'] = price
	break_out = False
	found = []
	for _filter in search_filters:
		# print(_filter)

		loop_index = 0
		axie_count = 0
		while True:
			try:
				search_parts_d = search_filters[_filter]['d']
				search_parts_r1 = search_filters[_filter]['r1']
				search_parts_r2 = search_filters[_filter]['r2']
				_class = search_filters[_filter]['class']
				price = search_filters[_filter]['price']
				# print(price)
				while True:
					data = start_marketquery_2(parts = search_parts_d, classes = _class, index = loop_index * 100, hp = search_filters[_filter]['hp'], morale = search_filters[_filter]['morale'], speed = search_filters[_filter]['speed'], 
					 skill = search_filters[_filter]['skill'], breedCount = search_filters[_filter]['breedCount'])

					# data = start_marketquery_2(parts = search_parts_d, index = loop_index * 100, classes = search_filters[_filter]['class'],
					#  hp = search_filters[_filter]['hp'], morale = search_filters[_filter]['morale'], speed = search_filters[_filter]['speed'], 
					#  skill = search_filters[_filter]['skill'], breedCount = search_filters[_filter]['breedCount'])
					time.sleep(sleep_time)
					total = data['data']['axies']['total']
					result = data['data']['axies']['results']
					# print(total)
					loop_index += 1
					for x in result:
						axie_count += 1
						genes = GetGenes256(x['genes'])
						d_parts = [genes.mouth.dId, genes.back.dId, genes.horn.dId, genes.tail.dId]
						d_parts = sorted(d_parts)
						r1_parts = [genes.mouth.r1Id, genes.back.r1Id, genes.horn.r1Id, genes.tail.r1Id]
						r1_parts = sorted(r1_parts)
						r2_parts = [genes.mouth.r2Id, genes.back.r2Id, genes.horn.r2Id, genes.tail.r2Id]
						r2_parts = sorted(r2_parts)
						print(genes.horn.dId)

						search_parts_d = sorted(search_parts_d)
						search_parts_r1 = sorted(search_parts_r1)
						search_parts_r2 = sorted(search_parts_r2)
						if x['owner'] == x['auction']['seller']:
							selling_price = convert_to_eth(x['auction']['currentPrice'])
							print(selling_price, price, selling_price > price)
							if selling_price > price:
								print("Over the budget")
								break_out = True
								break
							else:
								print(axie_count, "searched")
								# print("AXIE ID #", x['id'])
								# print(x['genes'])

								if search_parts_r1 != []:
									if search_parts_r2 != []:
										if r1_parts == search_parts_r1 and r2_parts == search_parts_r2:
											found.append(x)
											print(f"GET THIS! {selling_price}")
										else:
											print(f"R1 and R2 not the same. PRICE: {selling_price}")
									else:
										if r1_parts == search_parts_r1:
											found.append(x)
											print(f"GET THIS! {selling_price}")
										else:
											print(f"R1 not the same. PRICE: {selling_price}")

								else:
									found.append(x)
									print(f"GET THIS! {selling_price}")
						else:
							print("SOLD!")
					if break_out:
						#in preperation of the next loop
						break_out = False
						break

				break
			except ConnectionResetError:
				print("Connection Error, Retrying")

			except TypeError:
				print("Something went wrong. Retrying")

	wb = openpyxl.load_workbook("templates/email_excel_template.xlsx")
	ws = wb['D R1 R2']
	#codes to send email
	for _i, axie in enumerate(found):
		genes = GetGenes256(axie['genes'])
		max_row = ws.max_row + 1
		ws.cell(row = max_row, column = 1).value = axie['id']
		ws.cell(row = max_row, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['id']}/"
		ws.cell(row = max_row, column = 1).font = hyperlinked
		ws.cell(row = max_row, column = 2).value = axie['auction']['seller']
		ws.cell(row = max_row, column = 3).value = convert_to_eth(axie['auction']['currentPrice'])
		ws.cell(row = max_row, column = 3).number_format = "0.000"
		ws.cell(row = max_row, column = 3).alignment = Alignment(horizontal = "right")
		ws.cell(row = max_row, column = 4).value = axie['auction']['currentPriceUSD']
		ws.cell(row = max_row, column = 4).number_format = "0.000"
		ws.cell(row = max_row, column = 4).alignment = Alignment(horizontal = "right")
		ws.cell(row = max_row, column = 5).value = genes.cls.title()
		ws.cell(row = max_row, column = 6).value = axie['breedCount']
		ws.cell(row = max_row, column = 7).value = axie['stats']['hp']
		ws.cell(row = max_row, column = 8).value = axie['stats']['speed']
		ws.cell(row = max_row, column = 9).value = axie['stats']['skill']
		ws.cell(row = max_row, column = 10).value = axie['stats']['morale']
		ws.cell(row = max_row, column = 11).value = genes.horn.dName.title()
		ws.cell(row = max_row, column = 12).value = genes.mouth.dName.title()
		ws.cell(row = max_row, column = 13).value = genes.back.dName.title()
		ws.cell(row = max_row, column = 14).value = genes.tail.dName.title()
		ws.cell(row = max_row, column = 15).value = genes.horn.r1Name.title()
		ws.cell(row = max_row, column = 16).value = genes.mouth.r1Name.title()
		ws.cell(row = max_row, column = 17).value = genes.back.r1Name.title()
		ws.cell(row = max_row, column = 18).value = genes.tail.r1Name.title()
		ws.cell(row = max_row, column = 19).value = genes.horn.r2Name.title()
		ws.cell(row = max_row, column = 20).value = genes.mouth.r2Name.title()
		ws.cell(row = max_row, column = 21).value = genes.back.r2Name.title()
		ws.cell(row = max_row, column = 22).value = genes.tail.r2Name.title()

		#for color
		ws.cell(row = max_row, column = 11).font = Font(color = colors[parts_json[genes.horn.dId]['class']])
		ws.cell(row = max_row, column = 12).font = Font(color = colors[parts_json[genes.mouth.dId]['class']])
		ws.cell(row = max_row, column = 13).font = Font(color = colors[parts_json[genes.back.dId]['class']])
		ws.cell(row = max_row, column = 14).font = Font(color = colors[parts_json[genes.tail.dId]['class']])
		ws.cell(row = max_row, column = 15).font = Font(color = colors[parts_json[genes.horn.r1Id]['class']])
		ws.cell(row = max_row, column = 16).font = Font(color = colors[parts_json[genes.mouth.r1Id]['class']])
		ws.cell(row = max_row, column = 17).font = Font(color = colors[parts_json[genes.back.r1Id]['class']])
		ws.cell(row = max_row, column = 18).font = Font(color = colors[parts_json[genes.tail.r1Id]['class']])
		ws.cell(row = max_row, column = 19).font = Font(color = colors[parts_json[genes.horn.r2Id]['class']])
		ws.cell(row = max_row, column = 20).font = Font(color = colors[parts_json[genes.mouth.r2Id]['class']])
		ws.cell(row = max_row, column = 21).font = Font(color = colors[parts_json[genes.back.r2Id]['class']])
		ws.cell(row = max_row, column = 22).font = Font(color = colors[parts_json[genes.tail.r2Id]['class']])


	for i, row in enumerate(ws.iter_rows(min_row = 3)):
		if i % 2 == 0:
			for cell in row:
				cell.fill = PatternFill(fill_type="solid", start_color = "c8f4fa", end_color = "c8f4fa")

	wb.save("outputs/for_email.xlsx")


	if len(found) > 0:
		user = "ams.dummy.dev@gmail.com"
		password = "dummyamsdev"

		msg = EmailMessage()
		msg['Subject'] = "This is a test email2"
		msg['To'] = receiver
		msg['From'] = user
		msg.set_content(f"""
			Sent on {datetime.datetime.utcnow().strftime("%B, %d, %Y %I:%M:%S %p")}

			{json.dumps(f_json, indent = 4)}	
			""")

		with open("outputs/for_email.xlsx", "rb") as f:
			file_data = f.read()

		msg.add_attachment(file_data, maintype = "application", subtype = "xlsx", filename="filter_results.xlsx")

		with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
			server.login(user, password)
			server.send_message(msg)

		print("EMAIL SENT!")
	else:
		print("No match found, No email being sent.")

if __name__ == "__main__":
	delay_in_seconds = delay * 60
	def run_loop():
		run_program()
		threading.Timer(delay_in_seconds, run_loop).start()
		print("Running program")

	run_loop()