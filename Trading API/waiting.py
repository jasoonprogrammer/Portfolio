try:
	from selenium import webdriver
	from webdriver_manager.chrome import ChromeDriverManager
	from bs4 import BeautifulSoup
	from selenium.webdriver.chrome.options import Options
	from selenium.webdriver.common.by import By
	from selenium.webdriver.support.ui import WebDriverWait
	from selenium.webdriver.support import expected_conditions as EC
	import requests
	import time
	import openpyxl
	import datetime
	from openpyxl.styles import PatternFill, Font
	import json
	from selenium.common.exceptions import TimeoutException
except ModuleNotFoundError:
	print("Please run a pip install on the requirements.txt first")
	exit()

start_time = time.time()
chrome_options = Options()
wb = openpyxl.load_workbook("templates/Trading.xlsx")
sheet = wb.active
chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
chrome_options.add_argument("--headless")
installer = ChromeDriverManager().install()
driver = webdriver.Chrome(installer, options = chrome_options)
url = "https://dexscreener.com/new-pairs"
driver.get(url)
information = []
ethereum_networks = []
honeypot_json = {}
driver2 = webdriver.Chrome(installer, options = chrome_options)

while True:
	try:
		element = WebDriverWait(driver, 10).until(
			EC.presence_of_element_located((By.CLASS_NAME, "css-427d58"))
			)
		soup = BeautifulSoup(driver.page_source, "lxml")
		all_new = soup.find_all("a")
		for a in all_new[47:-1]:
			c = a.find_all("div", {"class": "css-1k4xub7"})
			if len(c) == 1:
				c = c[0].text
			elif len(c) == 2:
				c = c[1].text
			href = a.attrs['href']
			splitted_href = href.split("/")
			if "ethereum" in href:
				ethereum_networks.append((href, c))


		pairs = []
		api_url = 'https://api.dexscreener.io/latest/dex/pairs'

		for i, x in enumerate(ethereum_networks, start = 1):
			final_url = api_url + x[0]
			r = requests.get(final_url)
			r_dic = r.json()['pair']
			r_dic['age'] = x[1]
			information.append(r_dic)
			print(f"running API on index {i}")

		information = sorted(information, key = lambda x: x['pairCreatedAt'])
		information.reverse()

		for i, info in enumerate(information, start = 4):
			# print(info)
			sheet.cell(row = i, column = 1).value = info['chainId']
			sheet.cell(row = i, column = 2).value = info['dexId']
			sheet.cell(row = i, column = 3).value = info['url']
			sheet.cell(row = i, column = 4).value = info['pairAddress']
			sheet.cell(row = i, column = 4).hyperlink = "https://www.dextools.io/app/ether/pair-explorer/" + info['pairAddress']
			sheet.cell(row = i, column = 4).font = Font(color = "0000EE")
			sheet.cell(row = i, column = 5).value = info['baseToken']['address']
			sheet.cell(row = i, column = 6).value = info['baseToken']['name']
			sheet.cell(row = i, column = 7).value = info['baseToken']['symbol']
			sheet.cell(row = i, column = 8).value = info['quoteToken']['symbol']
			sheet.cell(row = i, column = 9).value = info['priceNative']
			sheet.cell(row = i, column = 10).value = info['priceUsd']
			try:
				sheet.cell(row = i, column = 11).value = info['txns']['h24']['buys']
				sheet.cell(row = i, column = 12).value = info['txns']['h24']['sells']
				sheet.cell(row = i, column = 19).value = info['volume']['h24']
				sheet.cell(row = i, column = 23).value = info['priceChange']['h24']

				sheet.cell(row = i, column = 13).value = info['txns']['h6']['buys']
				sheet.cell(row = i, column = 14).value = info['txns']['h6']['sells']
				sheet.cell(row = i, column = 20).value = info['volume']['h6']
				sheet.cell(row = i, column = 24).value = info['priceChange']['h6']

				sheet.cell(row = i, column = 15).value = info['txns']['h1']['buys']
				sheet.cell(row = i, column = 16).value = info['txns']['h1']['sells']
				sheet.cell(row = i, column = 21).value = info['volume']['h1']
				sheet.cell(row = i, column = 25).value = info['priceChange']['h1']

				sheet.cell(row = i, column = 17).value = info['txns']['m5']['buys']
				sheet.cell(row = i, column = 18).value = info['txns']['m5']['sells']
				sheet.cell(row = i, column = 22).value = info['volume']['m5']
				sheet.cell(row = i, column = 26).value = info['priceChange']['m5']

			except KeyError:
				#just leave the cell empty
				pass
			sheet.cell(row = i, column = 27).value = info['liquidity']['usd']
			sheet.cell(row = i, column = 27).number_format = "0"
			sheet.cell(row = i, column = 28).value = info['liquidity']['base']
			sheet.cell(row = i, column = 28).number_format = "0"
			sheet.cell(row = i, column = 29).value = info['liquidity']['quote']
			sheet.cell(row = i, column = 29).number_format = "0"
			sheet.cell(row = i, column = 30).value = info['fdv']
			sheet.cell(row = i, column = 30).number_format = "0"
			# sheet.cell(row = i, column = 31).value = info['pairCreatedAt']
			# sheet.cell(row = i, column = 31).number_format = "0"
			sheet.cell(row = i, column = 31).value = info['age']
			sheet.cell(row = i, column = 32).value = "Honeypot"
			sheet.cell(row = i, column = 32).hyperlink = "https://honeypot.is/ethereum?address=" + info['baseToken']['address']
			url2 = "https://honeypot.is/ethereum?address=" + info['baseToken']['address']
			driver2.get(url2)
			print("checking for honeypot...")
			while True:
				try:
					element2 = WebDriverWait(driver2, 5).until(
						EC.presence_of_element_located((By.ID, "token-info"))
						)
					soup2 = BeautifulSoup(driver2.page_source, "lxml")
					shtcoin = soup2.find(id = "shitcoin")
					msg = shtcoin.find("div", {"class": "compact"})
					honeypot_ = msg.find("div", {'class': "header"})
					print(info['baseToken']['name'], ":", honeypot_.text)
					is_honeypot = False if honeypot_.text == "Does not seem like a honeypot." else True

					if is_honeypot:
						sheet.cell(row = i, column = 32).value = "Honeypot"
						sheet.cell(row = i, column = 5).fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
						sheet.cell(row = i, column = 5).font = Font(color = "FFFFFF")
						sheet.cell(row = i, column = 32).font = Font(color = "FFFFFF")
						sheet.cell(row = i, column = 32).fill = PatternFill(start_color = "FF0000", end_color = "FF0000", fill_type = "solid")
					else:
						sheet.cell(row = i, column = 32).value = "Good"
						sheet.cell(row = i, column = 5).fill = PatternFill(start_color = "00FF00", end_color = "FF0000", fill_type = "solid")
						sheet.cell(row = i, column = 5).font = Font(color = "FFFFFF")
						sheet.cell(row = i, column = 32).font = Font(color = "FFFFFF")
						sheet.cell(row = i, column = 32).fill = PatternFill(start_color = "00FF00", end_color = "FF0000", fill_type = "solid")
						compact = msg.find_all("p")
						tax = None
						for x in compact:
							x_msg = x.prettify()
							if "Tax" in x_msg:
								tax = x_msg

						gases = tax.split("<br/>")
						gases = [x.strip() for x in gases]
						gases = [x.replace("<p>", "").replace("</p>", "") for x in gases]
						buy_tax = gases[0].split(":")[1].strip()
						sell_tax = gases[1].split(":")[1].strip()
						sheet.cell(row = i, column = 33).value = buy_tax.strip()
						sheet.cell(row = i, column = 34).value = sell_tax.strip()

						print("Buy Tax:", buy_tax.strip())
						print("Sell Tax:", sell_tax.strip())
					if is_honeypot:
						honeypot_json[info['baseToken']['name']] = {"honeypot": is_honeypot, "address": info['baseToken']['address'], "buyTax": None, "sellTax": None}
					else:
						honeypot_json[info['baseToken']['name']] = {"honeypot": is_honeypot, "address": info['baseToken']['address'], "buyTax": buy_tax, "sellTax": sell_tax}


					print("\n")
					break

				except TimeoutException:
					print("Error")
					time.sleep(3)





		d = datetime.datetime.now().strftime("%B_%d_%Y_%I_%M%p")
		with open(f"json_files/Honeypot_result_{d}.json", "w") as file:
			json.dump(honeypot_json, file)

		sheet.auto_filter.ref = "A3:AH3"
		wb.save(f"excel_files/results_{d}.xlsx")
		print("finished")
		print(f"Run time: {time.time() - start_time}")
		break


	except TimeoutException:
		print("error, reloading")
		time.sleep(3)
	
driver.close()
	
