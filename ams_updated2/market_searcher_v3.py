import requests
import time
import urllib
import openpyxl
import datetime
import json
import os
from openpyxl.styles import Alignment, Font, PatternFill
dirlist = os.listdir()
colors = {"beast": "FFB812", "aquatic": "00B8CE", "bird": "FF8BBD", "plant": "6CC000", "bug": "ff5341", "reptile": "B740CF", "dusk": "129092", "mech": "C6BDD4", "daw": "BECEFF"}


hyperlinked = Font(underline = "single", color="0C10F5")
if "market.json" not in dirlist:
	f = open("./market.json", 'w')
	f.write("[\n]")
	f.close()
	

def market_json_query(classes = [], parts = [], speed = [], morale = [], skill = [], hp = [], breedCount = [], pureness = []):
	f = open("./market.json", "r")
	jsonx = json.load(f)[:-1]
	queries = {'classes': classes, 'parts': parts, 'speed': speed, 'morale': morale, 'skill': skill, 'hp': hp, 'breedCount': breedCount, 'pureness': pureness}
	qs = [x['queries'] for x in jsonx]
	time_modified = os.path.getmtime("./market.json")
	dt = datetime.datetime.utcfromtimestamp(time_modified)
	time_delta = datetime.timedelta(minutes = 15)
	diff = datetime.datetime.utcnow() - dt
	past_15 = diff > time_delta
	print(diff)
	if queries not in qs or past_15:
		data = start_marketquery(classes = classes, parts = parts, speed = speed, morale = morale, skill = skill, hp = hp, breedCount = breedCount, pureness = pureness)
		d = {'queries': queries, "data": data}
		jsonx.append(d)
		with open("./market.json", "w") as outfile:
			json.dump(jsonx, outfile)
		return data
	else:
		for x in jsonx:
			if queries == x['queries']:
				return x['data']



def convert_to_eth(val):
	return int(val) / (10**18)

def marketquery(classes = [], parts = [], speed = [], morale = [], skill = [], hp = [], breedCount = [], pureness = [], stages = 4):
	classes = [x.title() for x in classes]
# data to be sent to api
	data ={
	"operationName": "GetAxieLatest",
	"variables": {
		"from": 0,
		"size": 10,
		"sort": "PriceAsc",
		"auctionType": "Sale",
		"criteria": {"classes": classes, "pureness": pureness, "parts": parts, "speed": speed, "morale": morale, "skill": skill, "hp": hp, "breedCount": breedCount, 'stages': stages}
		},
		"query": "query GetAxieLatest($auctionType: AuctionType, $criteria: AxieSearchCriteria, $from: Int, $sort: SortBy, $size: Int, $owner: String) {\n  axies(auctionType: $auctionType, criteria: $criteria, from: $from, sort: $sort, size: $size, owner: $owner) {\n    total\n    results {\n      ...AxieRowData\n      __typename\n    }\n    __typename\n  }\n}\n\nfragment AxieRowData on Axie {\n  id\n  image\n  class\n  name\n  genes\n  owner\n  class\n  stage\n  title\n  breedCount\n  level\n  parts {\n    ...AxiePart\n    __typename\n  }\n  stats {\n    ...AxieStats\n    __typename\n  }\n  auction {\n    ...AxieAuction\n    __typename\n  }\n  __typename\n}\n\nfragment AxiePart on AxiePart {\n  id\n  name\n  class\n  type\n  specialGenes\n  stage\n  abilities {\n    ...AxieCardAbility\n    __typename\n  }\n  __typename\n}\n\nfragment AxieCardAbility on AxieCardAbility {\n  id\n  name\n  attack\n  defense\n  energy\n  description\n  backgroundUrl\n  effectIconUrl\n  __typename\n}\n\nfragment AxieStats on AxieStats {\n  hp\n  speed\n  skill\n  morale\n  __typename\n}\n\nfragment AxieAuction on Auction {\n  startingPrice\n  endingPrice\n  startingTimestamp\n  endingTimestamp\n  duration\n  timeLeft\n  currentPrice\n  currentPriceUSD\n  suggestedPrice\n  seller\n  listingIndex\n  state\n  __typename\n}\n"
		}
	return data

def start_marketquery(build = "No Build Name", classes = [], parts = [], speed = [], morale = [], skill = [], hp = [], breedCount = [], pureness = []):
	API_ENDPOINT = "https://axieinfinity.com/graphql-server-v2/graphql"
	query_result = marketquery(classes = classes, parts = parts, speed = speed, morale = morale, skill = skill, hp = hp, pureness = pureness, stages = 4)
	r = requests.post(url = API_ENDPOINT, json = query_result)


	if r.status_code == 200:
		data = r.json()
		total = data['data']['axies']['total']
		print(total)
		if total <= 0:
			return {}
		else:
			axies = data['data']['axies']
			return data
	else:
		print(r)

def ownerquery(owner):
# data to be sent to api
	data ={
	"operationName": "GetAxieLatest",
	"variables": {
		"from": 0,
		"size": 100,
		"sort": "IdAsc",
		"auctionType": "All",
		"owner": owner
		},
		"query": "query GetAxieLatest($auctionType: AuctionType, $criteria: AxieSearchCriteria, $from: Int, $sort: SortBy, $size: Int, $owner: String) {\n  axies(auctionType: $auctionType, criteria: $criteria, from: $from, sort: $sort, size: $size, owner: $owner) {\n    total\n    results {\n      ...AxieRowData\n      __typename\n    }\n    __typename\n  }\n}\n\nfragment AxieRowData on Axie {\n  id\n  image\n  class\n  name\n  genes\n  owner\n  class\n  stage\n  title\n  breedCount\n  level\n  parts {\n    ...AxiePart\n    __typename\n  }\n  stats {\n    ...AxieStats\n    __typename\n  }\n  auction {\n    ...AxieAuction\n    __typename\n  }\n  __typename\n}\n\nfragment AxiePart on AxiePart {\n  id\n  name\n  class\n  type\n  specialGenes\n  stage\n  abilities {\n    ...AxieCardAbility\n    __typename\n  }\n  __typename\n}\n\nfragment AxieCardAbility on AxieCardAbility {\n  id\n  name\n  attack\n  defense\n  energy\n  description\n  backgroundUrl\n  effectIconUrl\n  __typename\n}\n\nfragment AxieStats on AxieStats {\n  hp\n  speed\n  skill\n  morale\n  __typename\n}\n\nfragment AxieAuction on Auction {\n  startingPrice\n  endingPrice\n  startingTimestamp\n  endingTimestamp\n  duration\n  timeLeft\n  currentPrice\n  currentPriceUSD\n  suggestedPrice\n  seller\n  listingIndex\n  state\n  __typename\n}\n"
		}
	return data

def start_ownerquery(owner):
	API_ENDPOINT = "https://axieinfinity.com/graphql-server-v2/graphql"
	query_result = ownerquery(owner)
	r = requests.post(url = API_ENDPOINT, json = query_result)

	if r.status_code == 200:
		data = r.json()
		total = data['data']['axies']['total']
		if total <= 0:
			return {}
		else:
			print("owner")
			axies = data['data']['axies']['results']

			return data
	else:
		print(r)


url_genes = "https://api.axie.technology/getgenes"
url_axies = "https://api.axie.technology/getaxies"



def axie_market_search():
	search_parts = ['mouth', 'horn', 'back', 'tail']
	wb = openpyxl.load_workbook("marketplace_template_v2.0.xlsx")
	ws = wb['Marketplace']
	axie_list = open("./axie.txt", "r")
	axies = [x.strip() for x in axie_list]
	joined = ",".join(axies)
	response = urllib.request.urlopen(f"{url_axies}/{joined}")
	axie_json = json.loads(response.read())
	for i, x in enumerate(axie_json):
		parts = []
		for p in x['parts']:
			if p['type'] == "Horn":
				parts.append(p['id'])
			if p['type'] == "Mouth":
				parts.append(p['id'])
			if p['type'] == "Back":
				parts.append(p['id'])
			if p['type'] == "Tail":
				parts.append(p['id'])
		classes = [x['class']]
		breedCount = [x['breedCount']]
		if x['breedCount'] != 0:
			breedCount.append(7)
		else:
			breedCount.append(0)
		speed = [x['stats']['speed'], 61]
		hp = [x['stats']['hp'], 61]
		morale = [x['stats']['morale'], 61]
		skill = [x['stats']['skill'], 61]
		result = market_json_query(parts = parts, classes = classes, breedCount = breedCount, hp = hp, speed = speed, morale = morale, skill = skill)
		result = result['data']['axies']
			
		if result != {}:
			m = ws.max_row + 1
			ws.cell(row = m, column = 1).value = x['story_id']
			ws.cell(row = m, column = 1).font = hyperlinked
			ws.cell(row = m, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{x['story_id']}/"
			ws.cell(row = m, column = 2).value = x['owner']
			ws.cell(row = m, column = 3).value = convert_to_eth(result['results'][0]['auction']['currentPrice'])
			ws.cell(row = m, column = 3).number_format = "0.000"
			total_5_usd = []
			total_5_eth = []
			total_10_usd = []
			total_10_eth = []
			for j, axie in enumerate(result['results']):
				if j < 5:
					total_5_usd.append(float(axie['auction']['currentPriceUSD']))
					total_5_eth.append(convert_to_eth(axie['auction']['currentPrice']))
				if j < 10:
					total_10_usd.append(float(axie['auction']['currentPriceUSD']))
					total_10_eth.append(convert_to_eth(axie['auction']['currentPrice']))
				c1 = 7 + (j * 2)
				c2 = 8 + (j * 2)
				ws.cell(row = m, column = c1).value = axie['id']
				ws.cell(row = m, column = c1).font = hyperlinked
				ws.cell(row = m, column = c1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['id']}/"
				ws.cell(row = m, column = c1).alignment = Alignment(horizontal = "right")
				ws.cell(row = m, column = c2).value = convert_to_eth(axie['auction']['currentPrice'])
				ws.cell(row = m, column = c2).number_format = "0.000"
				ws.cell(row = m, column = c2).alignment = Alignment(horizontal = "right")

			#getting the average
			ave_5_usd = sum(total_5_usd)/len(total_5_usd)
			ave_5_eth = sum(total_5_eth)/len(total_5_eth)
			ave_10_usd = sum(total_10_usd)/len(total_10_usd)
			ave_10_eth = sum(total_10_eth)/len(total_10_eth)
			ws.cell(row = m, column = 4).value = ave_10_eth
			ws.cell(row = m, column = 4).number_format = "0.000"
			ws.cell(row = m, column = 5).value = ave_10_usd
			ws.cell(row = m, column = 5).number_format = "0.000"
			ws.cell(row = m, column = 6).value = result['total']
			ws.cell(row = m, column = 27).value = x['class']
			for p in x['parts']:
				if p['type'] == "Horn":
					ws.cell(row = m, column = 28).value = p['name']
				if p['type'] == "Mouth":
					ws.cell(row = m, column = 29).value = p['name']
				if p['type'] == "Back":
					ws.cell(row = m, column = 30).value = p['name']
				if p['type'] == "Tail":
					ws.cell(row = m, column = 31).value = p['name']
			ws.cell(row = m, column = 32).value = datetime.datetime.now()
			ws.cell(row = m, column = 32).number_format = "MM/DD/YYYY HH:MM:SS"
			ws.cell(row = m, column = 33).value = time.tzname[0]
			for i, row in enumerate(ws.iter_rows(min_row = 2)):
				if i % 2 == 0:
					for cell in row:
						cell.fill = PatternFill(fill_type="solid", start_color = "c8f4fa", end_color = "c8f4fa")

	tday = datetime.datetime.today().strftime("%b_%d_%Y")
	dirlist = os.listdir()
	pile = [x for x in dirlist if f"axie_marketplace_{tday}" in x and "ronin" not in x]
	pile = [x.replace(f"axie_marketplace_{tday}_", "").replace(".xlsx", "") for x in pile]
	counts = [int(x) for x in pile]
	if len(counts) <= 0:
		d = 1
	else:
		d = max(counts) + 1
	wb.save(f"./axie_marketplace_{tday}_{d}.xlsx")


if __name__ == "__main__":
	print("\nSearch axies using?")
	print("1. axie.txt")
	print("2. ronin.txt")
	opt = input("")

	if int(opt) == 1:
		axie_market_search()

	elif int(opt) == 2:
		ronins = open("./ronin.txt", "r")
		ronin_ids = [x.strip().replace("ronin:", "0x") for x in ronins.readlines()]
		axie_list = []
		for x in ronin_ids:
			data = start_ownerquery(x)
			time.sleep(2)
			if data != {}:
				for ax in data['data']['axies']['results']:
					axie_list.append(ax)
		wb = openpyxl.load_workbook("ronin_marketplace_template_v2.0.xlsx")
		ws = wb['Marketplace']
		for i, x in enumerate(axie_list):
			parts = []
			for p in x['parts']:
				if p['type'] == "Horn":
					parts.append(p['id'])
				if p['type'] == "Mouth":
					parts.append(p['id'])
				if p['type'] == "Back":
					parts.append(p['id'])
				if p['type'] == "Tail":
					parts.append(p['id'])
			classes = [x['class']]
			classes = [x['class']]
			if x['class'] is not None:
				m = ws.max_row + 1
				breedCount = [x['breedCount']]
				if x['breedCount'] != 0:
					breedCount.append(7)
				else:
					breedCount.append(0)
				speed = [x['stats']['speed'], 61]
				hp = [x['stats']['hp'], 61]
				morale = [x['stats']['morale'], 61]
				skill = [x['stats']['skill'], 61]
				result = market_json_query(parts = parts, classes = classes, breedCount = breedCount, hp = hp, speed = speed, morale = morale, skill = skill)
				if result != {}:
					result = result['data']['axies']
					ws.cell(row = m, column = 1).value = x['id']
					ws.cell(row = m, column = 1).font = hyperlinked
					ws.cell(row = m, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{x['id']}/"
					ws.cell(row = m, column = 2).value = x['owner']
					total_5_usd = []
					total_5_eth = []
					total_10_usd = []
					total_10_eth = []

					if result != {}:
						for j, axie in enumerate(result['results']):
							print(axie['genes'])
							if j < 5:
								total_5_usd.append(float(axie['auction']['currentPriceUSD']))
								total_5_eth.append(convert_to_eth(axie['auction']['currentPrice']))
							if j < 10:
								total_10_usd.append(float(axie['auction']['currentPriceUSD']))
								total_10_eth.append(convert_to_eth(axie['auction']['currentPrice']))
							ws.cell(row = m, column = 3).value = convert_to_eth(result['results'][0]['auction']['currentPrice'])
							ws.cell(row = m, column = 3).number_format = "0.000"
							ws.cell(row = m, column = 3).alignment = Alignment(horizontal = "right")

							ws.cell(row = m, column = 6).value = result['total']
							c1 = 7 + (j * 2)
							c2 = 8 + (j * 2)
							ws.cell(row = m, column = c1).value = axie['id']
							ws.cell(row = m, column = c1).font = hyperlinked
							ws.cell(row = m, column = c1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['id']}/"
							ws.cell(row = m, column = c1).alignment = Alignment(horizontal = "right")
							ws.cell(row = m, column = c2).value = convert_to_eth(axie['auction']['currentPrice'])
							ws.cell(row = m, column = c2).number_format = "0.000"
							ws.cell(row = m, column = c2).alignment = Alignment(horizontal = "right")

						ave_5_usd = sum(total_5_usd)/len(total_5_usd)
						ave_5_eth = sum(total_5_eth)/len(total_5_eth)
						ave_10_usd = sum(total_10_usd)/len(total_10_usd)
						ave_10_eth = sum(total_10_eth)/len(total_10_eth)
						ws.cell(row = m, column = 4).value = ave_10_eth
						ws.cell(row = m, column = 4).number_format = "0.000"
						ws.cell(row = m, column = 5).value = ave_10_usd
						ws.cell(row = m, column = 5).number_format = "0.000"
				else:
					ws.cell(row = m, column = 1).value = x['id']
					ws.cell(row = m, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{x['id']}/"
					ws.cell(row = m, column = 1).font = hyperlinked				
					ws.cell(row = m, column = 2).value = x['owner']
					ws.cell(row = m, column = 3).value = 0
				ws.cell(row = m, column = 27).value = x['class']


				for p in x['parts']:
					if p['type'] == "Horn":
						ws.cell(row = m, column = 28).value = p['name']
					if p['type'] == "Mouth":
						ws.cell(row = m, column = 29).value = p['name']
					if p['type'] == "Back":
						ws.cell(row = m, column = 30).value = p['name']
					if p['type'] == "Tail":
						ws.cell(row = m, column = 31).value = p['name']
				ws.cell(row = m, column = 32).value = datetime.datetime.now()
				ws.cell(row = m, column = 32).number_format = "MM/DD/YYYY HH:MM:SS"
				ws.cell(row = m, column = 33).value = time.tzname[0]
				for i, row in enumerate(ws.iter_rows(min_row = 2)):
					if i % 2 == 0:
						for cell in row:
							cell.fill = PatternFill(fill_type="solid", start_color = "c8f4fa", end_color = "c8f4fa")


		tday = datetime.datetime.today().strftime("%b_%d_%Y")
		dirlist = os.listdir()
		pile = [x for x in dirlist if f"ronin_axie_marketplace_{tday}" in x]
		pile = [x.replace(f"ronin_axie_marketplace_{tday}_", "").replace(".xlsx", "") for x in pile]
		counts = [int(x) for x in pile]
		if len(counts) <= 0:
			d = 1
		else:
			d = max(counts) + 1
		wb.save(f"./ronin_axie_marketplace_{tday}_{d}.xlsx")






	else:
		print("Invalid Choice")
