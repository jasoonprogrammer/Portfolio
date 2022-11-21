import json
import urllib.request
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border
import os
import itertools
import time
from dependencies.mymodules import siblings, parents, request_data, start_marketquery, convert_to_eth
from dependencies.gene_class import GetGenes256

start_time = time.time()

url_axies = "https://api.axie.technology/getaxies"
url_validate = "https://api.axie.technology/invalidateaxie"

hyperlinked = Font(underline = "single", color="0C10F5")
colors = {"beast": "FFB812", "aquatic": "00B8CE", "bird": "FF8BBD", "plant": "6CC000", "bug": "ff5341", "reptile": "B740CF", "dusk": "129092", "mech": "C6BDD4", "daw": "BECEFF"}
cwd = os.getcwd().replace("\\", "/")
parts = open(f"{cwd}/json_files/parts.json")
parts_json = json.load(parts)
_id_json = {}
classes = {}
dirlist = os.listdir("outputs")
file_name = f"{cwd}/txt_files/axie.txt"
axie_ids = [x.strip() for x in open(file_name, "r").readlines()]
joined_axies = ",".join(axie_ids)
json_response = request_data(url = f"{url_axies}/{joined_axies},")[:-1]
for axie in json_response:
	if axie['genes'] == "0x0":
		request_data(url = f"{url_validate}/{axie['id']}")
json_response = request_data(url = f"{url_axies}/{joined_axies},")[:-1]

for axie in json_response:
	_id_json[axie['id']] = axie
	if axie['class'] in classes:
		classes[axie['class']].append(axie)
	else:
		classes[axie['class']] = [axie]


#check all possible breeding combinations where the pairs are not siblings or parents or themselves. and separated all axies by classes.
breeding_combinations = {}
for _classes, _axie in classes.items():
	_axie_list = sorted([x['id'] for x in _axie])
	for x in _axie_list:
		for y in _axie_list:
			axie_x = _id_json[x]
			axie_y = _id_json[y]
			if axie_x != axie_y and not parents(axie_x, axie_y) and not siblings(axie_x, axie_y):
				if _classes not in breeding_combinations:
					breeding_combinations[_classes] = [(x, y)]
				else:
					breeding_combinations[_classes].append((x, y))

###################################################



#display all combinations in excel file
wb = openpyxl.load_workbook(f"{cwd}/templates/breeding_template_v1.5.xlsx")
ws = wb['Breeding_Probability']
for _classes, _combinations in breeding_combinations.items():
	for _combination in _combinations:
		parent_1 = _id_json[_combination[0]]
		genes_1 = GetGenes256(parent_1['genes'])
		parent_2 = _id_json[_combination[1]]
		genes_2 = GetGenes256(parent_2['genes'])
		max_row = ws.max_row + 1
		ws.cell(row = max_row, column = 1).value = parent_1['id']
		ws.cell(row = max_row, column = 2).value = parent_2['id']
		ws.cell(row = max_row, column = 3).value = genes_1.purity
		ws.cell(row = max_row, column = 4).value = genes_2.purity
		probability = {	"eyes": genes_1.eyes + genes_2.eyes,
						"ears": genes_1.ears + genes_2.ears,
						"mouth": genes_1.mouth + genes_2.mouth,
						"back": genes_1.back + genes_2.back,
						"horn": genes_1.horn + genes_2.horn,
						"tail": genes_1.tail + genes_2.tail
						}
		bad_genes = {"eyes": 0, "ears": 0, "mouth": 0, "back": 0, "horn": 0, "tail": 0}
		for part, genes in probability.items():
			for k, v in genes.items():
				part = parts_json[k]
				if _classes.title() != part['class'].title():
					bad_genes[part['type']] += 1

		lengths = []
		for start, key in enumerate(probability):
			
			for i, gene in enumerate(probability[key].items()):
				_id, chance = gene
				ws.cell(row = max_row + i, column = 9 + start * 3).value = parts_json[_id]['name']
				ws.cell(row = max_row + i, column = 9 + start * 3).font = Font(color = colors[parts_json[_id]['class']])
				ws.cell(row = max_row + i, column = 10 + start * 3).font = Font(color = colors[parts_json[_id]['class']])
				ws.cell(row = max_row + i, column = 10 + start * 3).value = chance
				ws.cell(row = max_row + i, column = 10 + start * 3).number_format = "0.00"
				if i == len(probability[key]) - 1:
					lengths.append(i)

			ws.cell(row = max_row, column = 11 + start * 3).value = bad_genes[key]
		for i in range(1, 28):
			if i != 27:
				ws.cell(row = max_row + max(lengths), column = i).border = Border(bottom = Side(style = "thick"))
			else:
				ws.cell(row = max_row + max(lengths), column = i).border = Border(bottom = Side(style = "thick"), right = Side(style = "thick"))

		for i in range(max_row, max_row + max(lengths)):
			ws.cell(row = i, column = 27).border = Border(right = Side(style = "thick"))

################################


#checked all possible combinations of parents only from the same class from the pairings.
parent_combos = []
for _classes, _combinations in breeding_combinations.items():
	for _parents in _combinations:
		parent_combos.append([_parents[0], _parents[1]])
#####################




#loop through all possible combinations of pairings where there is no repeating parents.
combo_list = []
for i in range(len(parent_combos)):
	new_combo = list(parent_combos[i:] + parent_combos[:i])
	pairs = []
	looped = []
	for pair in new_combo:
		x, y = pair
		if x not in looped and y not in looped:
			pairs.append(sorted(pair))
			looped.append(x)
			looped.append(y)
	combo_list.append(pairs)
##########################


data_json = []
weights = {}
prices = []

#{"query": {"parts", "class"}, "data"}
queries = []

parents_result = []


#using all the possible combination of non-repeating parents
#checked their weight using the price of their highly possible offspring and store it in a list with their index in the 3rd list.
for i, _combo in enumerate(combo_list):
	total_weight = 0
	for _parents in _combo:
		result_list = []
		parents_list = [x['parents'] for x in parents_result]
		genes_parent1 = GetGenes256(_id_json[_parents[0]]['genes'])
		_class = genes_parent1.cls
		genes_parent2 = GetGenes256(_id_json[_parents[1]]['genes'])
		_class = _class.title()
		mouth = genes_parent1.mouth + genes_parent2.mouth
		back = genes_parent1.back + genes_parent2.back
		horn = genes_parent1.horn + genes_parent2.horn
		tail = genes_parent1.tail + genes_parent2.tail
		max_mouth = max(mouth, key = mouth.get)
		max_back = max(back, key = back.get)
		max_horn = max(horn, key = horn.get)
		max_tail = max(tail, key = tail.get)
		parts = [max_mouth, max_back, max_horn, max_tail]
		query = {"parts": parts, "class": _class}
		listed_query = [_iter['query'] for _iter in queries]
		if query in listed_query:
			index = listed_query.index(query)
			data = queries[index]['data']
			try:
				total = data['data']['axies']['total']
				axies = data['data']['axies']['results']
				for axie in axies:
					if axie['auction']['seller'] == axie['owner']:
						result_list.append(axie)
					if len(result_list) >= 10:
						break
				if len(result_list) != 0:
					queries.append({"query": query, "data": data})
					if _parents in parents_list:
						print("already in list")
					else:
						parents_result.append({"parents": _parents, "result": result_list})
					usd = [float(x['auction']['currentPriceUSD']) for x in result_list]
					if len(usd) > 0:
						average_usd = sum(usd) / len(usd)
					else:
						average_usd = 0
				else:
					axies = None
					queries.append({"query": query, "data": None})
					average_usd = 0
			except TypeError:
				axies = None
				queries.append({"query": query, "data": None})
				average_usd = 0





		else:
			data = start_marketquery(parts = parts, classes = [_class])
			time.sleep(1.5)
			total = data['data']['axies']['total']
			axies = data['data']['axies']['results']
			for axie in axies:
				if axie['auction']['seller'] == axie['owner']:
					result_list.append(axie)
				if len(result_list) >= 10:
					break
			if len(result_list) != 0:
				queries.append({"query": query, "data": data})
				if _parents in parents_list:
					print("already in list")
				else:
					parents_result.append({"parents": _parents, "result": result_list})
				usd = [float(x['auction']['currentPriceUSD']) for x in result_list]
				if len(usd) > 0:
					average_usd = sum(usd) / len(usd)
				else:
					average_usd = 0
			else:
				axies = None
				queries.append({"query": query, "data": None})
				average_usd = 0


		weight = mouth[max_mouth] * back[max_back] * horn[max_horn] * tail[max_tail] * .01 ** 4 * average_usd
		total_weight += weight

	weights[i] = total_weight

#######################

#getting the index of the combination with max weight
try:
	index_max_weight = max(weights, key = weights.get)
except ValueError as e:
	raise ValueError("No available combinations in the list of axies")

#the combination with highest weight
highest_weight = combo_list[index_max_weight]

ws2 = wb['Weight']
# looped though all that parents and listed it in the weights excel sheet.
for i, _parents in enumerate(highest_weight):
	_parent1 = _id_json[_parents[0]]
	_parent2 = _id_json[_parents[1]]
	_parent1_genes = GetGenes256(_parent1['genes'])
	_parent2_genes = GetGenes256(_parent2['genes'])
	_cls = _parent1_genes.cls
	mouth = _parent1_genes.mouth + _parent2_genes.mouth
	back = _parent1_genes.back + _parent2_genes.back
	horn = _parent1_genes.horn + _parent2_genes.horn
	tail = _parent1_genes.tail + _parent2_genes.tail
	max_mouth =  max(mouth, key = mouth.get)
	max_back =  max(back, key = back.get)
	max_horn =  max(horn, key = horn.get)
	max_tail =  max(tail, key = tail.get)
	parts = [max_mouth, max_back, max_horn, max_tail]
	usd = []
	for item in parents_result:
		if item['parents'] == _parents:
			axies = item['result']
			for axie in axies:
				usd.append(float(axie['auction']['currentPriceUSD']))
	try:
		average_usd = sum(usd) / len(usd)
	except ZeroDivisionError:
		#if no pair is found, default to 0
		average_usd = 0

	weight = mouth[max_mouth] * back[max_back] * horn[max_horn] * tail[max_tail] * 0.01 ** 4 * average_usd

	part_url = [f"&part={x}" for x in parts]
	part_url = "".join(part_url)
	class_url = f"?class={_cls.title()}"
	breed_count_url = "&breedCount=0"*2

	ws2.cell(row = 3 + i, column = 1).value = _parent1['id']
	ws2.cell(row = 3 + i, column = 2).value = _parent2['id']
	ws2.cell(row = 3 + i, column = 3).value = average_usd
	ws2.cell(row = 3 + i, column = 3).number_format = "0.000"
	ws2.cell(row = 3 + i, column = 4).value = parts_json[max_mouth]['name']
	ws2.cell(row = 3 + i, column = 5).value = mouth[max_mouth]
	ws2.cell(row = 3 + i, column = 5).number_format = "0.000"
	ws2.cell(row = 3 + i, column = 6).value = parts_json[max_back]['name']
	ws2.cell(row = 3 + i, column = 7).value = back[max_back]
	ws2.cell(row = 3 + i, column = 7).number_format = "0.000"
	ws2.cell(row = 3 + i, column = 8).value = parts_json[max_horn]['name']
	ws2.cell(row = 3 + i, column = 9).value = horn[max_horn]
	ws2.cell(row = 3 + i, column = 9).number_format = "0.000"
	ws2.cell(row = 3 + i, column = 10).value = parts_json[max_tail]['name']
	ws2.cell(row = 3 + i, column = 11).value = tail[max_tail]
	ws2.cell(row = 3 + i, column = 11).number_format = "0.000"
	ws2.cell(row = 3 + i, column = 12).value = weight
	ws2.cell(row = 3 + i, column = 12).number_format = "0.000"
	ws2.cell(row = 3 + i, column = 13).value = "Click Here"
	ws2.cell(row = 3 + i, column = 13).font = hyperlinked
	ws2.cell(row = 3 + i, column = 13).hyperlink = f"https://marketplace.axieinfinity.com/axie/{class_url}{part_url}{breed_count_url}"
######################################




# for i, _combo in enumerate(combo_list):
# 	total_weight = 0
# 	for _parents in _combo:
# 		genes_parent1 = GetGenes256(_id_json[_parents[0]]['genes'])
# 		_class = genes_parent1.cls
# 		genes_parent2 = GetGenes256(_id_json[_parents[1]]['genes'])
# 		_class = _class.title()
# 		mouth = genes_parent1.mouth + genes_parent2.mouth
# 		back = genes_parent1.back + genes_parent2.back
# 		horn = genes_parent1.horn + genes_parent2.horn
# 		tail = genes_parent1.tail + genes_parent2.tail
# 		max_mouth = max(mouth, key = mouth.get)
# 		max_back = max(back, key = back.get)
# 		max_horn = max(horn, key = horn.get)
# 		max_tail = max(tail, key = tail.get)
# 		parts = [max_mouth, max_back, max_horn, max_tail]
# 		query = {"class": _class.title(), "parts": parts}
# 		for _data in data_json:
# 			if _data['query'] not in queries:
# 				queries.append(_data['query'])
# 		if query in queries:
# 			for _data in data_json:
# 				if _data['query'] == query:
# 					data == _data['data']
# 		else:
# 			data = start_marketquery(parts = parts, classes = [_class])

# 			time.sleep(0.5)
# 		try:
# 			total = data['data']['axies']['total']
# 			axies = data['data']['axies']['results']
# 			usd = [round(float(x['auction']['currentPriceUSD']), 2) for x in axies]
# 			if len(usd) > 0:
# 				usd_average = sum(usd) / len(usd)
# 			else:
# 				usd_average = 0
# 			w = mouth[max_mouth] * back[max_back] * horn[max_horn] * tail[max_tail] * (0.01) ** 4 * usd_average
# 			data_json.append({'query': query, "data": data, "average": usd_average, "weight": w})
# 			total_weight += w
# 		except TypeError:
# 			print("No match found")
# 	weights[i] = total_weight

# max_weight_index = max(weights, key = weights.get)
# best_weight = combo_list[max_weight_index]
# ws2 = wb['Weight']

# for i, _parents in enumerate(best_weight):
# 	_parent1 = _id_json[_parents[0]]
# 	_parent2 = _id_json[_parents[1]]
# 	_genes_parent1 = GetGenes256(_parent1['genes'])
# 	_genes_parent2 = GetGenes256(_parent2['genes'])
# 	mouth = _genes_parent1.mouth + _genes_parent2.mouth
# 	back = _genes_parent1.back + _genes_parent2.back
# 	horn = _genes_parent1.horn + _genes_parent2.horn
# 	tail = _genes_parent1.tail + _genes_parent2.tail
# 	max_mouth = max(mouth, key = mouth.get)
# 	max_back = max(back, key = back.get)
# 	max_horn = max(horn, key = horn.get)
# 	max_tail = max(tail, key = tail.get)
# 	parts = [max_mouth, max_back, max_horn, max_tail]
# 	part_query = [f"&part={x}" for x in parts]
# 	part_query = "".join(part_query)
# 	class_query = f"?class={_parent1['class'].title()}"
# 	query = {"class": _parent1['class'].title(), "parts": parts}
# 	for _data in data_json:
# 		if _data['query'] == query:
# 			data == _data['data']
# 			average_usd = _data['average']
# 			w = mouth[max_mouth] * back[max_back] * horn[max_horn] * tail[max_tail] * (0.01) ** 4
# 			weight = average_usd * w

# 	ws2.cell(row = 3 + i, column = 1).value = _parent1['id']
# 	ws2.cell(row = 3 + i, column = 2).value = _parent2['id']
# 	ws2.cell(row = 3 + i, column = 3).value = average_usd
# 	ws2.cell(row = 3 + i, column = 4).value = parts_json[max_mouth]['name']
# 	ws2.cell(row = 3 + i, column = 5).value = mouth[max_mouth]
# 	ws2.cell(row = 3 + i, column = 6).value = parts_json[max_back]['name']
# 	ws2.cell(row = 3 + i, column = 7).value = back[max_back]
# 	ws2.cell(row = 3 + i, column = 8).value = parts_json[max_horn]['name']
# 	ws2.cell(row = 3 + i, column = 9).value = horn[max_horn]
# 	ws2.cell(row = 3 + i, column = 10).value = parts_json[max_tail]['name']
# 	ws2.cell(row = 3 + i, column = 11).value = tail[max_tail]
# 	ws2.cell(row = 3 + i, column = 12).value = weight
# 	ws2.cell(row = 3 + i, column = 12).number_format = "0.00"
# 	ws2.cell(row = 3 + i, column = 13).value = "Click Here"
# 	ws2.cell(row = 3 + i, column = 13).font = hyperlinked
# 	breed_count = f"&breedCount=0&breedCount=0"
# 	ws2.cell(row = 3 + i, column = 13).hyperlink = f"https://marketplace.axieinfinity.com/axie/{class_query}{part_query}{breed_count}"


wb.save(f"{cwd}/outputs/breeding_sheet.xlsx")
print(time.time() - start_time)
