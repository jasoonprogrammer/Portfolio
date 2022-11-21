import json
import urllib.request
import openpyxl
from openpyxl.styles import Font, Alignment, Side, Border
import os
from market_searcher_v2 import axie_market_search
import itertools
import time
from project import check_save_to_list
start_time = time.time()
f1 = open("./genes_list.json", "r")
f2 = open("./axies_list.json", "r")

colors = {"beast": "FFB812", "aquatic": "00B8CE", "bird": "FF8BBD", "plant": "6CC000", "bug": "ff5341", "reptile": "B740CF"}


# data_genes = json.load(f1)
# data_axies = json.load(f2)

axies = open("./axie.txt", "r")
axie_list = [x.strip() for x in axies.readlines()]
axies = ",".join(axie_list)

# url_genes = "https://api.axie.technology/getgenes"
# url_axies = "https://api.axie.technology/getaxies"


# response_genes = urllib.request.urlopen(f"{url_genes}/{axies}")
# response_axies = urllib.request.urlopen(f"{url_axies}/{axies}")
query = check_save_to_list(axie_list)
data_genes = query['genes']
data_axies = query['details']
breed = []

for x1 in data_axies:
	for x2 in data_axies:
		if (
			x1['id'] == x2['id'] or x1['id'] == x2['sireId'] or x1['id'] == x2['matronId'] or
			x1['sireId'] == x2['id'] or x1['sireId'] == x2['sireId'] or x1['sireId'] == x2['matronId'] or
			x1['matronId'] == x2['id'] or x1['matronId'] == x2['sireId'] or x1['matronId'] == x2['matronId']
			):
			# print("Axies Can't Breed (Related)")
			pass
		else:
			
			parents = [int(x1['id']), int(x2['id'])]
			parent_listed = [x['parents'] for x in breed]
			if parents not in parent_listed:
				axie_1_genes = ""
				axie_1_genes = ""

				for x in data_genes:
					if x['axieId'] == x1['id']:
						axie_1_genes = x
					if x['axieId'] == x2['id']:
						axie_2_genes = x

				# axie_1_genes = json.loads(urllib.request.urlopen(f"{url_genes}/{x1['id']}").read())
				# axie_2_genes = json.loads(urllib.request.urlopen(f"{url_genes}/{x2['id']}").read())

				pattern_probability = {}
				color_probability = {}
				eyes_probability = {}
				ears_probability = {}
				mouth_probability = {}
				horn_probability = {}
				back_probability = {}
				tail_probability = {}
				purity_parts = ['mouth', 'back', 'horn', "eyes", "ears", "tail"]

				for i, x in enumerate(data_genes):
					if x['axieId'] == x1['id']:
						axie_1_genes = x
					if x['axieId'] == x2['id']:
						axie_2_genes = x





				#calculate for purity
				axie_1_purity = 100
				axie_2_purity = 100
				for x in purity_parts:
					axie_1_class = x1['class'].lower()
					axie_2_class = x2['class'].lower()

					for key, value in axie_1_genes[x].items():
						if key == "d":
							if value['class'] != axie_1_class:
								axie_1_purity -= 12.5
						elif key == "r1":
							if value['class'] != axie_1_class:
								axie_1_purity -= 3
						elif key == "r2":
							if value['class'] != axie_1_class:
								axie_1_purity -= 1

					for key, value in axie_2_genes[x].items():
						if key == "d":
							if value['class'] != axie_2_class:
								axie_2_purity -= 12.5
						elif key == "r1":
							if value['class'] != axie_2_class:
								axie_2_purity -= 3
						elif key == "r2":
							if value['class'] != axie_2_class:
								axie_2_purity -= 1

				

				purity = []
				for p in parents:
					if p == int(x1['id']):
						purity.clear()
						purity.append((axie_1_purity, x1['class']))
						purity.append((axie_2_purity, x2['class']))
					else:
						purity.clear()
						purity.append((axie_2_purity, x2['class']))
						purity.append((axie_1_purity, x2['class']))



				#pattern
				for key, value in axie_1_genes['pattern'].items():
					v = str(value)
					if key == "d":
						if v in pattern_probability:
							pattern_probability[v] += 37.5
						else:
							pattern_probability[v] = 37.5
					elif key == "r1":
						if v in pattern_probability:
							pattern_probability[v] += 9.375
						else:
							pattern_probability[v] = 9.375
					elif key == "r2":
						if v in pattern_probability:
							pattern_probability[v] += 3.125
						else:
							pattern_probability[v] = 3.125

				for key, value in axie_2_genes['pattern'].items():
					v = str(value)
					if key == "d":
						if v in pattern_probability:
							pattern_probability[v] += 37.5
						else:
							pattern_probability[v] = 37.5
					elif key == "r1":
						if v in pattern_probability:
							pattern_probability[v] += 9.375
						else:
							pattern_probability[v] = 9.375
					elif key == "r2":
						if v in pattern_probability:
							pattern_probability[v] += 3.125
						else:
							pattern_probability[v] = 3.125

				#color
				for key, value in axie_1_genes['color'].items():
					v = str(value)
					if key == "d":
						if v in color_probability:
							color_probability[v] += 37.5
						else:
							color_probability[v] = 37.5
					elif key == "r1":
						if v in color_probability:
							color_probability[v] += 9.375
						else:
							color_probability[v] = 9.375
					elif key == "r2":
						if v in color_probability:
							color_probability[v] += 3.125
						else:
							color_probability[v] = 3.125

				for key, value in axie_2_genes['color'].items():
					v = str(value)
					if key == "d":
						if v in color_probability:
							color_probability[v] += 37.5
						else:
							color_probability[v] = 37.5
					elif key == "r1":
						if v in color_probability:
							color_probability[v] += 9.375
						else:
							color_probability[v] = 9.375
					elif key == "r2":
						if v in color_probability:
							color_probability[v] += 3.125
						else:
							color_probability[v] = 3.125


				#ears
				for key, value in axie_1_genes['ears'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['name'] in ears_probability:
								ears_probability[value['name']]['probability'] += 37.5
							else:
								ears_probability[value['name']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['name'] in ears_probability:
								ears_probability[value['name']]['probability'] += 9.375
							else:
								ears_probability[value['name']] = {"class": value['class'], "probability": 9.375}
								ears_probability[value['name']]['probability'] = 9.375
						elif key == "r2":
							if value['name'] in ears_probability:
								ears_probability[value['name']]['probability'] += 3.125
							else:
								ears_probability[value['name']] = {"class": value['class'], "probability": 3.125}
								ears_probability[value['name']]['probability'] = 3.125

				for key, value in axie_2_genes['ears'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['name'] in ears_probability:
								ears_probability[value['name']]['probability'] += 37.5
							else:
								ears_probability[value['name']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['name'] in ears_probability:
								ears_probability[value['name']]['probability'] += 9.375
							else:
								ears_probability[value['name']] = {"class": value['class'], "probability": 9.375}
								ears_probability[value['name']]['probability'] = 9.375
						elif key == "r2":
							if value['name'] in ears_probability:
								ears_probability[value['name']]['probability'] += 3.125
							else:
								ears_probability[value['name']] = {"class": value['class'], "probability": 3.125}
								ears_probability[value['name']]['probability'] = 3.125

					#eyes
				for key, value in axie_1_genes['eyes'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['name'] in eyes_probability:
								eyes_probability[value['name']]['probability'] += 37.5
							else:
								eyes_probability[value['name']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['name'] in eyes_probability:
								eyes_probability[value['name']]['probability'] += 9.375
							else:
								eyes_probability[value['name']] = {"class": value['class'], "probability": 9.375}
								eyes_probability[value['name']]['probability'] = 9.375
						elif key == "r2":
							if value['name'] in eyes_probability:
								eyes_probability[value['name']]['probability'] += 3.125
							else:
								eyes_probability[value['name']] = {"class": value['class'], "probability": 3.125}
								eyes_probability[value['name']]['probability'] = 3.125

				for key, value in axie_2_genes['eyes'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['name'] in eyes_probability:
								eyes_probability[value['name']]['probability'] += 37.5
							else:
								eyes_probability[value['name']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['name'] in eyes_probability:
								eyes_probability[value['name']]['probability'] += 9.375
							else:
								eyes_probability[value['name']] = {"class": value['class'], "probability": 9.375}
								eyes_probability[value['name']]['probability'] = 9.375
						elif key == "r2":
							if value['name'] in eyes_probability:
								eyes_probability[value['name']]['probability'] += 3.125
							else:
								eyes_probability[value['name']] = {"class": value['class'], "probability": 3.125}
								eyes_probability[value['name']]['probability'] = 3.125




					#mouth
				for key, value in axie_1_genes['mouth'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['name'] in mouth_probability:
								mouth_probability[value['name']]['probability'] += 37.5
							else:
								mouth_probability[value['name']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['name'] in mouth_probability:
								mouth_probability[value['name']]['probability'] += 9.375
							else:
								mouth_probability[value['name']] = {"class": value['class'], "probability": 9.375}
								mouth_probability[value['name']]['probability'] = 9.375
						elif key == "r2":
							if value['name'] in mouth_probability:
								mouth_probability[value['name']]['probability'] += 3.125
							else:
								mouth_probability[value['name']] = {"class": value['class'], "probability": 3.125}
								mouth_probability[value['name']]['probability'] = 3.125

				for key, value in axie_2_genes['mouth'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['name'] in mouth_probability:
								mouth_probability[value['name']]['probability'] += 37.5
							else:
								mouth_probability[value['name']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['name'] in mouth_probability:
								mouth_probability[value['name']]['probability'] += 9.375
							else:
								mouth_probability[value['name']] = {"class": value['class'], "probability": 9.375}
								mouth_probability[value['name']]['probability'] = 9.375
						elif key == "r2":
							if value['name'] in mouth_probability:
								mouth_probability[value['name']]['probability'] += 3.125
							else:
								mouth_probability[value['name']] = {"class": value['class'], "probability": 3.125}
								mouth_probability[value['name']]['probability'] = 3.125






				#horn
				for key, value in axie_1_genes['horn'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['partId'] in horn_probability:
								horn_probability[value['partId']]['probability'] += 37.5
							else:
								horn_probability[value['partId']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['partId'] in horn_probability:
								horn_probability[value['partId']]['probability'] += 9.375
							else:
								horn_probability[value['partId']] = {"class": value['class'], "probability": 9.375}
								horn_probability[value['partId']]['probability'] = 9.375
						elif key == "r2":
							if value['partId'] in horn_probability:
								horn_probability[value['partId']]['probability'] += 3.125
							else:
								horn_probability[value['partId']] = {"class": value['class'], "probability": 3.125}
								horn_probability[value['partId']]['probability'] = 3.125

				for key, value in axie_2_genes['horn'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['partId'] in horn_probability:
								horn_probability[value['partId']]['probability'] += 37.5
							else:
								horn_probability[value['partId']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['partId'] in horn_probability:
								horn_probability[value['partId']]['probability'] += 9.375
							else:
								horn_probability[value['partId']] = {"class": value['class'], "probability": 9.375}
								horn_probability[value['partId']]['probability'] = 9.375
						elif key == "r2":
							if value['partId'] in horn_probability:
								horn_probability[value['partId']]['probability'] += 3.125
							else:
								horn_probability[value['partId']] = {"class": value['class'], "probability": 3.125}
								horn_probability[value['partId']]['probability'] = 3.125



				#back
				for key, value in axie_1_genes['back'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['partId'] in back_probability:
								back_probability[value['partId']]['probability'] += 37.5
							else:
								back_probability[value['partId']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['partId'] in back_probability:
								back_probability[value['partId']]['probability'] += 9.375
							else:
								back_probability[value['partId']] = {"class": value['class'], "probability": 9.375}
								back_probability[value['partId']]['probability'] = 9.375
						elif key == "r2":
							if value['partId'] in back_probability:
								back_probability[value['partId']]['probability'] += 3.125
							else:
								back_probability[value['partId']] = {"class": value['class'], "probability": 3.125}
								back_probability[value['partId']]['probability'] = 3.125

				for key, value in axie_2_genes['back'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['partId'] in back_probability:
								back_probability[value['partId']]['probability'] += 37.5
							else:
								back_probability[value['partId']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['partId'] in back_probability:
								back_probability[value['partId']]['probability'] += 9.375
							else:
								back_probability[value['partId']] = {"class": value['class'], "probability": 9.375}
								back_probability[value['partId']]['probability'] = 9.375
						elif key == "r2":
							if value['partId'] in back_probability:
								back_probability[value['partId']]['probability'] += 3.125
							else:
								back_probability[value['partId']] = {"class": value['class'], "probability": 3.125}
								back_probability[value['partId']]['probability'] = 3.125


				#tail
				for key, value in axie_1_genes['tail'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['partId'] in tail_probability:
								tail_probability[value['partId']]['probability'] += 37.5
							else:
								tail_probability[value['partId']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['partId'] in tail_probability:
								tail_probability[value['partId']]['probability'] += 9.375
							else:
								tail_probability[value['partId']] = {"class": value['class'], "probability": 9.375}
								tail_probability[value['partId']]['probability'] = 9.375
						elif key == "r2":
							if value['partId'] in tail_probability:
								tail_probability[value['partId']]['probability'] += 3.125
							else:
								tail_probability[value['partId']] = {"class": value['class'], "probability": 3.125}
								tail_probability[value['partId']]['probability'] = 3.125

				for key, value in axie_2_genes['tail'].items():
					if isinstance(value, dict):
						if key == "d":
							if value['partId'] in tail_probability:
								tail_probability[value['partId']]['probability'] += 37.5
							else:
								tail_probability[value['partId']] = {"class": value['class'], "probability": 37.5}
						elif key == "r1":
							if value['partId'] in tail_probability:
								tail_probability[value['partId']]['probability'] += 9.375
							else:
								tail_probability[value['partId']] = {"class": value['class'], "probability": 9.375}
								tail_probability[value['partId']]['probability'] = 9.375
						elif key == "r2":
							if value['partId'] in tail_probability:
								tail_probability[value['partId']]['probability'] += 3.125
							else:
								tail_probability[value['partId']] = {"class": value['class'], "probability": 3.125}
								tail_probability[value['partId']]['probability'] = 3.125


				x = {
				"purity": purity, "parents": parents, "probability": {"color": color_probability, "pattern": pattern_probability, "eyes": eyes_probability, "ears": ears_probability,
				"mouth": mouth_probability, "horn": horn_probability, "back": back_probability, "tail": tail_probability}
				}
				breed.append(x)

				not_needed = ["color", "ears", "eyes", "pattern"]
				#sort the probability by highest
				for p in x['probability']:
					if p not in not_needed:
						item = x['probability'][p]
						z = sorted(item.items(), reverse = True, key = lambda tup: (tup[1]["probability"]))
						z = dict(z)
						x['probability'][p] = z


with open("./breed_list.json", "w") as outfile:
	json.dump(breed, outfile)

wb = openpyxl.load_workbook("./breeding_template_v1.5.xlsx")
ws = wb['Breeding_Probability']
for x in breed:
	parent_id = x['parents'][0]
	parent_class = x['purity'][0][1].lower()

	lengths = [len(x['probability'][b]) for b in x['probability']]
	m = max(lengths)
	r = ws.max_row + 1
	for i in range(m):
		ws.cell(row = r+i, column = 1).value = x['parents'][0]
		ws.cell(row = r+i, column = 2).value = x['parents'][1]
		ws.cell(row = r+i, column = 3).value = f"Purity {x['purity'][0][0]}"
		ws.cell(row = r+i, column = 4).value = f"Purity {x['purity'][1][0]}"
		ws.cell(row = r+i, column = 1).alignment = Alignment(horizontal = "center")
		ws.cell(row = r+i, column = 2).alignment = Alignment(horizontal = "center")
		ws.cell(row = r+i, column = 3).alignment = Alignment(horizontal = "center")
		ws.cell(row = r+i, column = 4).alignment = Alignment(horizontal = "center")
		ws.cell(row = r+i, column = 1).border = Border(left= Side(border_style = "thick", color = "000000"))

	for i, a in enumerate(x['probability']['color']):
		ws.cell(row = r + i, column = 5).value = a.upper()
		ws.cell(row = r + i, column = 5).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 6).value = x['probability']['color'][a]
		ws.cell(row = r + i, column = 6).number_format = "#.00"

	for i, a in enumerate(x['probability']['pattern']):
		ws.cell(row = r + i, column = 7).value = str(a.upper())
		ws.cell(row = r + i, column = 7).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 8).value = x['probability']['pattern'][a]
		ws.cell(row = r + i, column = 8).number_format = "#.00"

	bad_eyes = 0
	for i, a in enumerate(x['probability']['eyes']):
		ws.cell(row = r + i, column = 9).value = a.upper()
		ws.cell(row = r + i, column = 9).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 9).font = Font(color = colors[x['probability']['eyes'][a]['class']])
		ws.cell(row = r + i, column = 10).value = x['probability']['eyes'][a]['probability']
		if parent_class != x['probability']['eyes'][a]['class']:
			bad_eyes += 1
		ws.cell(row = r + i, column = 10).number_format = "#.00"

	ws.cell(row = r, column = 11).value = bad_eyes

	bad_ears = 0
	for i, a in enumerate(x['probability']['ears']):
		ws.cell(row = r + i, column = 12).value = a.upper()
		ws.cell(row = r + i, column = 12).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 12).font = Font(color = colors[x['probability']['ears'][a]['class']])
		ws.cell(row = r + i, column = 13).value = x['probability']['ears'][a]['probability']
		if parent_class != x['probability']['ears'][a]['class']:
			bad_ears += 1
		ws.cell(row = r + i, column = 13).number_format = "#.00"

	ws.cell(row = r, column = 14).value = bad_ears

	bad_mouth = 0
	for i, a in enumerate(x['probability']['mouth']):
		ws.cell(row = r + i, column = 15).value = a.upper()
		ws.cell(row = r + i, column = 15).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 15).font = Font(color = colors[x['probability']['mouth'][a]['class']])
		ws.cell(row = r + i, column = 16).value = x['probability']['mouth'][a]['probability']
		ws.cell(row = r + i, column = 16).number_format = "#.00"

		if parent_class != x['probability']['mouth'][a]['class']:
			bad_mouth += 1
	ws.cell(row = r, column = 17).value = bad_mouth

	bad_back = 0
	for i, a in enumerate(x['probability']['back']):
		ws.cell(row = r + i, column = 18).value = a.upper()
		ws.cell(row = r + i, column = 18).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 18).font = Font(color = colors[x['probability']['back'][a]['class']])
		ws.cell(row = r + i, column = 19).value = x['probability']['back'][a]['probability']
		ws.cell(row = r + i, column = 19).number_format = "#.00"
		if parent_class != x['probability']['back'][a]['class']:
			bad_back += 1
	ws.cell(row = r, column = 20).value = bad_back

	bad_horn = 0
	for i, a in enumerate(x['probability']['horn']):
		ws.cell(row = r + i, column = 21).value = a.upper()
		ws.cell(row = r + i, column = 21).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 21).font = Font(color = colors[x['probability']['horn'][a]['class']])
		ws.cell(row = r + i, column = 22).value = x['probability']['horn'][a]['probability']
		ws.cell(row = r + i, column = 22).number_format = "#.00"
		if parent_class != x['probability']['horn'][a]['class']:
			bad_horn += 1
	ws.cell(row = r, column = 23).value = bad_horn

	bad_tail = 0
	for i, a in enumerate(x['probability']['tail']):
		ws.cell(row = r + i, column = 24).value = a.upper()
		ws.cell(row = r + i, column = 24).alignment = Alignment(horizontal = "center")
		ws.cell(row = r + i, column = 24).font = Font(color = colors[x['probability']['tail'][a]['class']])
		ws.cell(row = r + i, column = 25).value = x['probability']['tail'][a]['probability']
		ws.cell(row = r + i, column = 25).number_format = "#.00"
		if parent_class != x['probability']['tail'][a]['class']:
			bad_tail += 1
	ws.cell(row = r, column = 26).value = bad_tail

	for i in range(m):
		ws.cell(row = r + i, column = 27).value = i + 1
		ws.cell(row = r + i, column = 27).border = Border(right = Side(border_style="thick", color = "000000"))

	for i in range(1, 28):
		if i == 1:
			ws.cell(row = ws.max_row, column = i).border = Border(bottom = Side(border_style = "thick", color = "000000"), left = Side(border_style = "thick", color = "000000"))
		elif i == 27:
			ws.cell(row = ws.max_row, column = i).border = Border(bottom = Side(border_style = "thick", color = "000000"), right = Side(border_style = "thick", color = "000000"))
		else:
			ws.cell(row = ws.max_row, column = i).border = Border(bottom = Side(border_style = "thick", color = "000000"))

	# for i, a in enumerate(x['probability']['']):
	# 	ws.cell(row = r + i, column = 3).value = a.upper()
	# 	ws.cell(row = r + i, column = 4).value = x['probability']['color'][a]

	# for i, a in enumerate(x['probability']['color']):
	# 	ws.cell(row = r + i, column = 3).value = a.upper()
	# 	ws.cell(row = r + i, column = 4).value = x['probability']['color'][a]

	# for i, a in enumerate(x['probability']['color']):
	# 	ws.cell(row = r + i, column = 3).value = a.upper()
	# 	ws.cell(row = r + i, column = 4).value = x['probability']['color'][a]

	# for i, a in enumerate(x['probability']['color']):
	# 	ws.cell(row = r + i, column = 3).value = a.upper()
	# 	ws.cell(row = r + i, column = 4).value = x['probability']['color'][a]

	# for i, a in enumerate(x['probability']['color']):
	# 	ws.cell(row = r + i, column = 3).value = a.upper()
	# 	ws.cell(row = r + i, column = 4).value = x['probability']['color'][a]

#array of parent

#get all average values of axie in axie.txt to prevent rerunning


axie_list = set(axie_list)
axie_list = [int(x) for x in axie_list]
parent_combos = []
# for row in ws.iter_rows(min_row = 3):
# 	parent_combos.append([row[0].value, row[1].value])

for x in breed:
	parent_combos.append([x['parents'][0], x['parents'][1]])

combo_list = []
for i in range(len(parent_combos)):
	new_combo = list(parent_combos[i:] + parent_combos[:i])
	pairs = []
	looped = []
	for pair in new_combo:
		x, y = pair
		if x not in looped and y not in looped:
			pairs.append(sorted(pair))
			looped.append(x)
			looped.append(y)
	combo_list.append(pairs)


#to prevent rerunning the same xlsx file
breeding_json = []
done = []
for row in ws.iter_rows(min_row = 3):
	pair = [row[0].value, row[1].value]
	pair.sort()
	if pair not in done:
		weights = [row[9].value, row[12].value, row[15].value, row[18].value]
		breeding_json.append({"parents": pair, "weights": weights})
	done.append(pair)

#getting price
dirlist = os.listdir()
dirlist = [x for x in dirlist if "axie_marketplace" in x and "ronin" not in x]
if len(dirlist) <= 0:
	axie_market_search()
	dirlist = os.listdir()
	dirlist = [x for x in dirlist if "axie_marketplace" in x and "ronin" not in x]
f_name = dirlist[-1]

check_price = openpyxl.load_workbook(f_name)
ws = check_price["Marketplace"]
axie_price = {}
for x in ws.iter_rows(min_row = 2):
	axie_price[int(x[0].value)] = x[4].value

weight_comparison = {}

for i, combo in enumerate(combo_list):

	for pair in combo:
		combo_weight = []
		x, y = pair
		for breed_pair in breeding_json:
			if pair == breed_pair['parents']:
				w = breed_pair['weights']
				w_total = w[0] * w[1] * w[2] * w[3] * (.01 ** 4)
				p1_weight = w_total * axie_price[breed_pair['parents'][0]]
				p2_weight = w_total * axie_price[breed_pair['parents'][1]]
				total_weight = p1_weight + p2_weight
				combo_weight.append(total_weight)
		weight_comparison[i] = sum(combo_weight)

max_key = max(weight_comparison, key = weight_comparison.get)
# print(max_key)
# print(weight_comparison[max_key])

ws2 = wb['Weight']
for i, pair in enumerate(combo_list[max_key], start = 3):
	ws2.cell(row = i, column = 1).value = pair[0]
	ws2.cell(row = i, column = 2).value = pair[1]
	for breed_pair in breeding_json:
		if pair == breed_pair['parents']:
			w = breed_pair['weights']
			ws2.cell(row = i, column = 3).value = axie_price[breed_pair['parents'][0]]
			ws2.cell(row = i, column = 3).number_format = "0.000"
			ws2.cell(row = i, column = 4).value = axie_price[breed_pair['parents'][1]]
			ws2.cell(row = i, column = 4).number_format = "0.000"
			ws2.cell(row = i, column = 5).value = w[0]
			ws2.cell(row = i, column = 6).value = w[1]
			ws2.cell(row = i, column = 7).value = w[2]
			ws2.cell(row = i, column = 8).value = w[3]
			w_total = w[0] * w[1] * w[2] * w[3] * (.01 ** 4)
			p1_weight = w_total * axie_price[breed_pair['parents'][0]]
			p2_weight = w_total * axie_price[breed_pair['parents'][1]]
			ws2.cell(row = i, column = 9).value = p1_weight
			ws2.cell(row = i, column = 9).number_format = "0.000"
			ws2.cell(row = i, column = 10).value = p2_weight
			ws2.cell(row = i, column = 10).number_format = "0.000"
			ws2.cell(row = i, column = 11).value = p1_weight + p2_weight
			ws2.cell(row = i, column = 11).number_format = "0.000"


	# done = []
	# for row in ws.iter_rows(min_row = 3):
	# 	pair = [row[0].value, row[1].value]
	# 	pair.sort()
	# 	if pair in combo and pair not in done:
	# 		done.append(pair)
	# 		print(row[9].value, row[12].value, row[15].value, row[18].value)



# parents_on_list = []
# for x in ws.iter_rows(min_row = 3):

# 	#expanding them to check if on of the parents in the list
# 	all_parents = []
# 	for p in parents_on_list:
# 		all_parents += p

# 	parents = []
# 	parents.append(x[0].value)
# 	parents.append(x[1].value)
# 	parents.append(x[15].value)
# 	parents.append(x[18].value)
# 	parents.append(x[21].value)
# 	parents.append(x[24].value)

# 	#if either parents not in list. append them to parent list
# 	if x[0].value not in all_parents and x[1].value not in all_parents:
# 		parents_on_list.append(parents)





# check_price = openpyxl.load_workbook(f_name)
# ws = check_price["Marketplace"]
# axie_price = {}
# for x in ws.iter_rows(min_row = 2):
# 	axie_price[x[0].value] = x[4].value




# ws2 = wb['Weight']
# for parent in parents_on_list:
# 	m = ws2.max_row + 1
# 	ws2.cell(row = m, column = 1).value = parent[0]
# 	ws2.cell(row = m, column = 2).value = parent[1]
# 	ws2.cell(row = m, column = 3).value = axie_price[str(parent[0])]
# 	ws2.cell(row = m, column = 4).value = axie_price[str(parent[1])]
# 	ws2.cell(row = m, column = 5).value = parent[2]
# 	ws2.cell(row = m, column = 6).value = parent[3]
# 	ws2.cell(row = m, column = 7).value = parent[4]
# 	ws2.cell(row = m, column = 8).value = parent[5]
# 	w1 = parent[2] * parent[2] * parent[2] * parent[2] * float(axie_price[str(parent[0])]) * (.01 ** 4) #(.01 ** 4) is for the x/100 of 4 parts
# 	w2 = parent[2] * parent[2] * parent[2] * parent[2] * float(axie_price[str(parent[1])]) * (.01 ** 4)
# 	ws2.cell(row = m, column = 9).value = w1
# 	ws2.cell(row = m, column = 9).number_format = "0.000"
# 	ws2.cell(row = m, column = 10).value = w2
# 	ws2.cell(row = m, column = 10).number_format = "0.000"




wb.save("./breeding_sheet.xlsx")
print(time.time() - start_time)

