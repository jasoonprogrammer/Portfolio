import requests
import time
import urllib
import openpyxl
import datetime
import json
import os
from openpyxl.styles import Alignment, Font, PatternFill
from dependencies.gene_class import *
from dependencies.mymodules import Stat, index_query, start_indexquery
import math
cwd = os.getcwd().replace("\\", "/")
start_time = time.time()
dirlist = os.listdir("json_files")
colors = {"beast": "FFB812", "aquatic": "00B8CE", "bird": "FF8BBD", "plant": "6CC000", "bug": "ff5341", "reptile": "B740CF", "dusk": "129092", "mech": "C6BDD4", "daw": "BECEFF"}



hyperlinked = Font(underline = "single", color="0C10F5")
if "market.json" not in dirlist:
	print("not found")
	f = open("json_files/market.json", 'w')
	f.write("[\n]")
	f.close()
else:
	print("found")


time_modified = os.path.getmtime("json_files/market.json")
dt = datetime.datetime.utcfromtimestamp(time_modified)
time_delta = datetime.timedelta(minutes = 5)
diff = datetime.datetime.utcnow() - dt
past_15 = diff > time_delta

if not past_15:
	f = open(cwd + "/json_files/market.json", "r")
	jsonx = json.load(f)
else:
	print("outdated market.json")
	jsonx = []


def market_json_query(classes = [], parts = [], speed = [], morale = [], skill = [], hp = [], breedCount = [], pureness = [], past_15 = True):
	queries = {'classes': classes, 'parts': parts, 'speed': speed, 'morale': morale, 'skill': skill, 'hp': hp, 'breedCount': breedCount, 'pureness': pureness}
	qs = [x['queries'] for x in jsonx]
	if queries not in qs:
		print("from api")
		data = start_marketquery(classes = classes, parts = parts, speed = speed, morale = morale, skill = skill, hp = hp, breedCount = breedCount, pureness = pureness)
		d = {'queries': queries, "data": data}
		jsonx.append(d)
		return data
	else:
		print("from json")
		for x in jsonx:
			if queries == x['queries']:
				return x['data']



def convert_to_eth(val):
	return int(val) / (10**18)

def marketquery(classes = [], parts = [], speed = [], morale = [], skill = [], hp = [], breedCount = [], pureness = [], stages = 4):
	classes = [x.title() for x in classes]
# data to be sent to api
	data ={
	"operationName": "GetAxieLatest",
	"variables": {
		"from": 0,
		"size": 30,
		"sort": "PriceAsc",
		"auctionType": "Sale",
		"criteria": {"classes": classes, "pureness": pureness, "parts": parts, "speed": speed, "morale": morale, "skill": skill, "hp": hp, "breedCount": breedCount, 'stages': stages}
		},
		"query": "query GetAxieLatest($auctionType: AuctionType, $criteria: AxieSearchCriteria, $from: Int, $sort: SortBy, $size: Int, $owner: String) {\n  axies(auctionType: $auctionType, criteria: $criteria, from: $from, sort: $sort, size: $size, owner: $owner) {\n    total\n    results {\n      ...AxieRowData\n      __typename\n    }\n    __typename\n  }\n}\n\nfragment AxieRowData on Axie {\n  id\n  image\n  class\n  name\n  genes\n  owner\n  class\n  stage\n  title\n  breedCount\n  level\n  parts {\n    ...AxiePart\n    __typename\n  }\n  stats {\n    ...AxieStats\n    __typename\n  }\n  auction {\n    ...AxieAuction\n    __typename\n  }\n  __typename\n}\n\nfragment AxiePart on AxiePart {\n  id\n  name\n  class\n  type\n  specialGenes\n  stage\n  abilities {\n    ...AxieCardAbility\n    __typename\n  }\n  __typename\n}\n\nfragment AxieCardAbility on AxieCardAbility {\n  id\n  name\n  attack\n  defense\n  energy\n  description\n  backgroundUrl\n  effectIconUrl\n  __typename\n}\n\nfragment AxieStats on AxieStats {\n  hp\n  speed\n  skill\n  morale\n  __typename\n}\n\nfragment AxieAuction on Auction {\n  startingPrice\n  endingPrice\n  startingTimestamp\n  endingTimestamp\n  duration\n  timeLeft\n  currentPrice\n  currentPriceUSD\n  suggestedPrice\n  seller\n  listingIndex\n  state\n  __typename\n}\n"
		}
	return data

def start_marketquery(build = "No Build Name", classes = [], parts = [], speed = [], morale = [], skill = [], hp = [], breedCount = [], pureness = []):
	while True:
		try:
			API_ENDPOINT = "https://axieinfinity.com/graphql-server-v2/graphql"
			query_result = marketquery(classes = classes, parts = parts, speed = speed, morale = morale, skill = skill, hp = hp, pureness = pureness, stages = 4, breedCount = breedCount)
			r = requests.post(url = API_ENDPOINT, json = query_result)


			if r.status_code == 200:
				data = r.json()
				total = data['data']['axies']['total']
				if total <= 0:
					return {}
					break
				else:
					axies = data['data']['axies']
					return data
					break
			else:
				print(r)
				break
		except TypeError:
			time.sleep(1.5)

def ownerquery(owner):
# data to be sent to api
	data ={
	"operationName": "GetAxieLatest",
	"variables": {
		"from": 0,
		"size": 100,
		"sort": "IdAsc",
		"auctionType": "All",
		"owner": owner,
		"criteria": {"stages": 4}
		},
		"query": "query GetAxieLatest($auctionType: AuctionType, $criteria: AxieSearchCriteria, $from: Int, $sort: SortBy, $size: Int, $owner: String) {\n  axies(auctionType: $auctionType, criteria: $criteria, from: $from, sort: $sort, size: $size, owner: $owner) {\n    total\n    results {\n      ...AxieRowData\n      __typename\n    }\n    __typename\n  }\n}\n\nfragment AxieRowData on Axie {\n  id\n  image\n  class\n  name\n  genes\n  owner\n  class\n  stage\n  title\n  breedCount\n  level\n  parts {\n    ...AxiePart\n    __typename\n  }\n  stats {\n    ...AxieStats\n    __typename\n  }\n  auction {\n    ...AxieAuction\n    __typename\n  }\n  __typename\n}\n\nfragment AxiePart on AxiePart {\n  id\n  name\n  class\n  type\n  specialGenes\n  stage\n  abilities {\n    ...AxieCardAbility\n    __typename\n  }\n  __typename\n}\n\nfragment AxieCardAbility on AxieCardAbility {\n  id\n  name\n  attack\n  defense\n  energy\n  description\n  backgroundUrl\n  effectIconUrl\n  __typename\n}\n\nfragment AxieStats on AxieStats {\n  hp\n  speed\n  skill\n  morale\n  __typename\n}\n\nfragment AxieAuction on Auction {\n  startingPrice\n  endingPrice\n  startingTimestamp\n  endingTimestamp\n  duration\n  timeLeft\n  currentPrice\n  currentPriceUSD\n  suggestedPrice\n  seller\n  listingIndex\n  state\n  __typename\n}\n"
		}
	return data

def start_ownerquery(owner):
	while True:
		try:
			API_ENDPOINT = "https://axieinfinity.com/graphql-server-v2/graphql"
			query_result = ownerquery(owner)
			r = requests.post(url = API_ENDPOINT, json = query_result)

			if r.status_code == 200:
				data = r.json()
				total = data['data']['axies']['total']
				if total <= 0:
					return {}
					break
				else:
					axies = data['data']['axies']['results']

					return data
					break
			else:
				print(r)
				break
		except TypeError:
			time.sleep(1.5)



url_genes = "https://api.axie.technology/getgenes"
url_axies = "https://api.axie.technology/getaxies"
url_validate = "https://api.axie.technology/invalidateaxie"

def axie_market_search(axie_list = [], wb = None, d_r1 = True):
	wb = wb
	search_parts = ['mouth', 'horn', 'back', 'tail']
	if axie_list == []:
		axie_list = open(cwd + "/txt_files/axie.txt", "r")
	axies = [x.strip() for x in axie_list]
	joined = ",".join(axies)
	response = urllib.request.urlopen(f"{url_axies}/{joined},")
	axie_json = json.loads(response.read())[:-1]
	for axie in axie_json:
		try:
			print("checking")
			genes = GetGenes256(axie['genes'])
		except KeyError:
			_id = axie['story_id']
			response = urllib.request.urlopen(f"{url_validate}/{_id}")
			time.sleep(1.5)


	response = urllib.request.urlopen(f"{url_axies}/{joined},")
	axie_json = json.loads(response.read())[:-1]

	axie_json = [x for x in axie_json if x['stage'] == 4]
	r1_list = {}
	pair_price = {}
	for i, x in enumerate(axie_json):
		if x['class'] is not None:
			genes = GetGenes256(x['genes'])
			classes = x['class']
			m = ws.max_row + 1
			breedCount = [0, 0]
			past_15 = diff > time_delta
			temp_list = []
			speed = [x['stats']['speed'], 61]
			morale = [x['stats']['morale'], 61]
			skill = [x['stats']['skill'], 61]
			hp = [x['stats']['hp'], 61]
			##check if d == r1
			main_axie_better_mouth = genes.mouth.dId == genes.mouth.r1Id
			main_axie_better_horn = genes.horn.dId == genes.horn.r1Id
			main_axie_better_back = genes.back.dId == genes.back.r1Id
			main_axie_better_tail = genes.tail.dId == genes.tail.r1Id
			result = start_indexquery(classes = classes, index = 0, speed = speed, morale = morale, skill = skill, hp = hp, parts = [genes.mouth.dId, genes.back.dId, genes.tail.dId, genes.horn.dId])

			if x['breedCount'] == 0 and result != {}:
				total_count = result['data']['axies']['total']
				loop_count = math.ceil(total_count / 30) + 1
				for index in range(loop_count):
					break_out = False
					c_index = index * 30
					result = start_indexquery(classes = classes, index = c_index, speed = speed, morale = morale, skill = skill, hp = hp, parts = [genes.mouth.dId, genes.back.dId, genes.tail.dId, genes.horn.dId])
					time.sleep(3)
					for r in result['data']['axies']['results']:
						if r['auction']['seller'] == r['owner'] and (main_axie_better_back or main_axie_better_horn or main_axie_better_mouth or main_axie_better_tail):
							match_genes = GetGenes256(r['genes'])
							match_genes = GetGenes256(r['genes'])
							##check if d == r1
							match_mouth = match_genes.mouth.dId == match_genes.mouth.r1Id
							match_horn = match_genes.horn.dId == match_genes.horn.r1Id
							match_back = match_genes.back.dId == match_genes.back.r1Id
							match_tail = match_genes.tail.dId == match_genes.tail.r1Id
							if (
								#if match's d == r1 is true, and main axie's d == r1 is also false; then returns true
								#if main axie's d== r1 and match's d != r1, return false
								#True = 1, False = 0 for reference
								(match_mouth >= main_axie_better_mouth) and
								(match_horn >= main_axie_better_horn) and
								(match_back >= main_axie_better_back) and
								(match_tail >= main_axie_better_tail)
								):
								print(r['id'])
								temp_list.append(r)
								#check temp list size
								if len(temp_list) >= 10:
									print("temp_list full")
									r1_list[x['id']] = temp_list
									break_out = True
									break
							else:
								#do nothing
								pass
					#break out of loop
					if break_out:
						print("out of loop")
						break

					print("list didn't reach 10")
					r1_list[x['id']] = temp_list

	time.sleep(3)
	for _id, axies in r1_list.items():
		max_row = ws.max_row + 1
		eths = [convert_to_eth(x['auction']['currentPrice']) for x in axies]
		if len(eths) > 0:
			average_eth = sum(eths) / len(eths)
			min_eth = min(eths)
		else:
			average_eth = ""
			min_eth = ""
		pair_price[_id] = {"min": min_eth, "average": average_eth}


	for i, x in enumerate(axie_json):
		genes = GetGenes256(x['genes'])
		parts = [genes.mouth.dId, genes.tail.dId, genes.back.dId, genes.horn.dId]
		classes = [genes.cls]
		breedCount = [x['breedCount']]
		if x['breedCount'] != 0:
			breedCount.append(7)
		else:
			breedCount.append(0)
		speed = [x['stats']['speed'], 61]
		hp = [x['stats']['hp'], 61]
		morale = [x['stats']['morale'], 61]
		skill = [x['stats']['skill'], 61]
		result = market_json_query(parts = parts, classes = classes, hp = hp, speed = speed, morale = morale, skill = skill, past_15 = past_15)

		if result != {}:
			r = []
			result = result['data']['axies']
			for z in result['results']:
				if z['auction']['seller'] == z['owner']:
					print("Not sold")
					r.append(z)
					if len(r) >= 10:
						break
			m = ws.max_row + 1
			if x['auction'] is None:
				ws.cell(row = m, column = 1).value = x['story_id']
				ws.cell(row = m, column = 1).font = hyperlinked
			else:
				ws.cell(row = m, column = 1).value = x['story_id']
				ws.cell(row = m, column = 1).font = Font(strike = True, color="0C10F5")
			ws.cell(row = m, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{x['story_id']}/"
			ws.cell(row = m, column = 2).value = x['owner']
			ws.cell(row = m, column = 3).value = convert_to_eth(r[0]['auction']['currentPrice'])
			ws.cell(row = m, column = 3).number_format = "0.000"
			if d_r1 and len(r1_list) > 0:
				ws.cell(row = m, column = 8).value = pair_price[x['story_id']]['min']
				ws.cell(row = m, column = 8).number_format = "0.000"
				ws.cell(row = m, column = 9).value = pair_price[x['story_id']]['average']
				ws.cell(row = m, column = 9).number_format = "0.000"
			total_5_usd = []
			total_5_eth = []
			total_10_usd = []
			total_10_eth = []
			for j, axie in enumerate(r):
				print(f"ID: {axie['id']}, GENE ID: {axie['genes']}")
				if j < 5:
					total_5_usd.append(float(axie['auction']['currentPriceUSD']))
					total_5_eth.append(convert_to_eth(axie['auction']['currentPrice']))
				if j < 10:
					total_10_usd.append(float(axie['auction']['currentPriceUSD']))
					total_10_eth.append(convert_to_eth(axie['auction']['currentPrice']))
				c1 = 10 + (j * 2)
				c2 = 11 + (j * 2)
				ws.cell(row = m, column = c1).value = axie['id']
				ws.cell(row = m, column = c1).font = hyperlinked
				ws.cell(row = m, column = c1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['id']}/"
				ws.cell(row = m, column = c1).alignment = Alignment(horizontal = "right")
				ws.cell(row = m, column = c2).value = convert_to_eth(axie['auction']['currentPrice'])
				ws.cell(row = m, column = c2).number_format = "0.000"
				ws.cell(row = m, column = c2).alignment = Alignment(horizontal = "right")

			#getting the average
			ave_5_usd = sum(total_5_usd)/len(total_5_usd)
			ave_5_eth = sum(total_5_eth)/len(total_5_eth)
			ave_10_usd = sum(total_10_usd)/len(total_10_usd)
			ave_10_eth = sum(total_10_eth)/len(total_10_eth)
			ws.cell(row = m, column = 4).value = ave_10_eth
			ws.cell(row = m, column = 4).number_format = "0.000"
			ws.cell(row = m, column = 5).value = ave_10_usd
			ws.cell(row = m, column = 5).number_format = "0.000"
			ws.cell(row = m, column = 6).value = result['total']
			ws.cell(row = m, column = 7).value = x['breedCount']
			ws.cell(row = m, column = 30).value = x['class']
			for p in x['parts']:
				if p['type'] == "Horn":
					ws.cell(row = m, column = 31).value = p['name']
				if p['type'] == "Mouth":
					ws.cell(row = m, column = 32).value = p['name']
				if p['type'] == "Back":
					ws.cell(row = m, column = 33).value = p['name']
				if p['type'] == "Tail":
					ws.cell(row = m, column = 34).value = p['name']
			ws.cell(row = m, column = 35).value = datetime.datetime.now()
			ws.cell(row = m, column = 35).number_format = "MM/DD/YYYY HH:MM:SS"
			ws.cell(row = m, column = 36).value = time.tzname[0]
			for i, row in enumerate(ws.iter_rows(min_row = 2)):
				if i % 2 == 0:
					for cell in row:
						cell.fill = PatternFill(fill_type="solid", start_color = "c8f4fa", end_color = "c8f4fa")
		else:
			print(x['id'])

	

if __name__ == "__main__":
	print("\nSearch axies using?")
	print("1. axie.txt")
	print("2. ronin.txt")
	opt = input("")

	if int(opt) == 1:
		while True:
			prompt_response = input("Run D = R1? (Y/N)\n")
			if prompt_response.lower().strip() == "n" or prompt_response.strip() == "":
				d_r1 = False
				wb = openpyxl.load_workbook("./templates/hidden_marketplace_template_v3.0.xlsx")
				ws = wb['Marketplace']
				break
			elif prompt_response.lower().strip() == "y":
				d_r1 = True
				wb = openpyxl.load_workbook("./templates/marketplace_template_v3.0.xlsx")
				ws = wb['Marketplace']
				break
			else:
				print("Invalid Choice")
		axie_market_search(wb = wb, d_r1 = d_r1)

	elif int(opt) == 2:
		while True:
			prompt_response = input("Run D = R1? (Y/N)\n")
			if prompt_response.lower().strip() == "n" or prompt_response.strip() == "":
				d_r1 = False
				wb = openpyxl.load_workbook("./templates/hidden_marketplace_template_v3.0.xlsx")
				ws = wb['Marketplace']
				break
			elif prompt_response.lower().strip() == "y":
				d_r1 = True
				wb = openpyxl.load_workbook("./templates/marketplace_template_v3.0.xlsx")
				ws = wb['Marketplace']
				break
			else:
				print("Invalid Choice")
		ronins = open(cwd + "/txt_files/ronin.txt", "r")
		ronin_ids = [x.strip().replace("ronin:", "0x") for x in ronins.readlines()]
		axie_list = []
		for x in ronin_ids:
			data = start_ownerquery(x)
			time.sleep(2)
			if data != {}:
				for ax in data['data']['axies']['results']:
					axie_list.append(ax['id'])
		axie_market_search(axie_list, wb = wb, d_r1 = d_r1)






	else:
		print("Invalid Choice")
tday = datetime.datetime.today().strftime("%m_%d_%Y")
dirlist = os.listdir("outputs")
pile = [x for x in dirlist if f"axie_marketplace_{tday}" in x]
pile = [x.replace(f"axie_marketplace_{tday}_", "").replace(".xlsx", "").replace("~$", "") for x in pile]
counts = [int(x) for x in pile]
if len(counts) <= 0:
	d = 1
else:
	d = max(counts) + 1
wb.save(f"outputs/axie_marketplace_{tday}_{d}.xlsx")
with open("json_files/market.json", "w") as outfile:
	json.dump(jsonx, outfile)
print(f"Runtime {time.time() - start_time}")
