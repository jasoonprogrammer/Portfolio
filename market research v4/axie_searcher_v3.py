import os, openpyxl, datetime
from dependencies.market_search_modules import ownerquery, start_ownerquery, start_marketquery
import math, time, json
from dependencies.gene_class import GetGenes256
from dependencies.mymodules import convert_to_eth
import urllib
from openpyxl.styles import Font, Alignment, PatternFill
from socket import error as SocketError
maxsleep = 1.5
maxretries = 3
start_time = datetime.datetime.now()
hyperlinked = Font(underline = "single", color="0C10F5")
striked = Font(strike = True, color = "0C10F5")


time_modified = os.path.getmtime("json_files/market_json.json")
dt = datetime.datetime.utcfromtimestamp(time_modified)
time_delta = datetime.timedelta(minutes = 5)
diff = datetime.datetime.utcnow() - dt
past_5 = diff > time_delta

url_axies = "https://api.axie.technology/getaxies"
url_validate = "https://api.axie.technology/invalidateaxie"

dirlist = os.listdir("json_files")
if "market_json.json" not in dirlist:
	print("market_json not found")
	f = open("json_files/market_json.json", 'w')
	f.write("[\n]")
	f.close()
else:
	print("market_json found")

if "market_json_d_r1.json" not in dirlist:
	print("market_json_d_r1 not found")
	f2 = open("json_files/market_json_d_r1.json", 'w')
	f2.write("[\n]")
	f2.close()
else:
	print("market_json_d_r1 found")

if not past_5:
	print("Loading JSON files.")
	r_query_file = open("json_files/market_json.json", "r")
	r_query = json.load(r_query_file)
	r1_query_file = open("json_files/market_json_d_r1.json", "r")
	r1_query = json.load(r1_query_file)
else:
	print("Old json data, overwriting files.")
	r_query = []
	r1_query = []


axie_txt = open("txt_files/axie.txt")
axie_list = [x.strip() for x in axie_txt.readlines()]
joined = ",".join(axie_list)
response = urllib.request.urlopen(f"{url_axies}/{joined},")
time.sleep(maxsleep)
json_response = json.loads(response.read())[:-1]
axie_txt.close()
for axie in json_response:
	if axie['genes'] == "0x0":
		revalidate = urllib.request.urlopen(f"{url_validate}/{axie['id']}")
		time.sleep(maxsleep)

#retrieve updated data
response = urllib.request.urlopen(f"{url_axies}/{joined},")
time.sleep(1.5)
json_response = json.loads(response.read())[:-1]

axie_list = [x for x in json_response if x['stage'] == 4]

while True:
	print("Run D = R1 = R2? (Y/N)\n")
	choice = input("")
	if choice.lower() == "n" or choice == "":
		d_r1 = False
		wb = openpyxl.load_workbook("templates/hidden_marketplace_template_v3.0.xlsx")
		ws = wb['Marketplace']
		break
	elif choice.lower() == "y":
		d_r1 = True
		wb = openpyxl.load_workbook("templates/marketplace_template_v3.0.xlsx")
		ws = wb['Marketplace']
		break
	else:
		print("Invalid Choice")


for axie in axie_list:
	stats = axie['stats']
	hp = [stats['hp'], 61]
	morale = [stats['morale'], 61]
	speed = [stats['speed'], 61]
	skill = [stats['skill'], 61]
	genes = GetGenes256(axie['genes'])
	parts = [genes.mouth.dId, genes.back.dId, genes.horn.dId, genes.tail.dId]

	main_axie_better_mouth = genes.mouth.dId == genes.mouth.r1Id
	main_axie_better_horn = genes.horn.dId == genes.horn.r1Id
	main_axie_better_back = genes.back.dId == genes.back.r1Id
	main_axie_better_tail = genes.tail.dId == genes.tail.r1Id

	#check for perfect genes
	main_axie_perfect_mouth = genes.mouth.dId == genes.mouth.r1Id == genes.mouth.r2Id
	main_axie_perfect_horn = genes.horn.dId == genes.horn.r1Id == genes.horn.r2Id
	main_axie_perfect_back = genes.back.dId == genes.back.r1Id == genes.back.r2Id
	main_axie_perfect_tail = genes.tail.dId == genes.tail.r1Id == genes.tail.r2Id
	print(main_axie_perfect_mouth, "PERFECT MOUTH MAIN")
	print(main_axie_perfect_horn, "PERFECT HORN MAIN")
	print(main_axie_perfect_back, "PERFECT BACK MAIN")
	print(main_axie_perfect_tail, "PERFECT TAIL MAIN")

	query = {"class": axie['class'], "parts": parts, "stats": stats}
	r_response = list(filter(lambda q: q['query'] == query, r_query))
	r1_response = list(filter(lambda q: q['query'] == query, r1_query))
	retries = 0

	r1_list = []
	r = []
	while True:
		market_response = start_marketquery(classes = [axie['class']], parts = parts, hp = hp, morale = morale, skill = skill, speed = speed)
		total = market_response['data']['axies']['total']
		if r_response != []:
			r = r_response[0]['data']
			print("Aquiring data from json")
			break
		else:
			print("Aquiring data from API")
			try:
				if total != 0:
					time.sleep(maxsleep)
					print("result found")
					break
				else:
					if retries >= maxretries:
						print("Max retries! No axies found")
						break
					retries += 1
					print(f"No axies found, rerunning results: Retries: {retries}")
			except ConnectionResetError as e:
				time.sleep(1)
				print("Server ConnectionError. Retrying")
			except SocketError as e:
				time.sleep(1)
				print("Server ConnectionError. Retrying")

			except TypeError:
				time.sleep(1)
				total = 0
				retries += 1
				print(f"No axies found, rerunning results: Retries: {retries}")

				if retries >= 5:
					print("Max retries! No axies found")

					break
	if d_r1 and axie['breedCount'] == 0:

		if r1_response != []:
			r1_list = r1_response[0]['data']
			print("Aquiring d_r1 data from json")
		else:
			market_response = start_marketquery(classes = [axie['class']], parts = parts, hp = hp, morale = morale, skill = skill, speed = speed)
			total = market_response['data']['axies']['total']

			loops = math.ceil(total / 30)
			break_out = False
			for i in range(0, loops):
				if r1_response != []:
					r1_list = r1_response[0]['data']
					break

				retries_2 = 0

				while True:
					try:
						if total != 0:
							time.sleep(maxsleep)
							break
						else:
							if retries_2 >= 5:
								print("Max retries! No axies found")
								break
							retries_2 += 1
							print(f"No axies found, rerunning results: Retries: {retries_2}")
					except ConnectionResetError as e:
						time.sleep(maxsleep)
						print("Server ConnectionError. Retrying")
					except SocketError as e:
						time.sleep(maxsleep)
						print("Server ConnectionError. Retrying")

					except TypeError:
						time.sleep(maxsleep)
						retries_2 += 1
						print(f"No axies found, rerunning results: Retries: {retries_2}")

						if retries_2 >= 5:
							total = 0
							print("Max retries! No axies found")

							break
				while True:
					try:
						qq = start_marketquery(classes = [axie['class']], parts = parts, hp = hp, morale = morale, skill = skill, speed = speed, index = i * 30)
						axies = qq['data']['axies']['results']
						total = qq['data']['axies']['total']
						print(total)
						time.sleep(maxsleep)
						break
					except TypeError:
						print("No axies found, retrying")
						market_retry += 1
						if market_retry >= maxretries:
							print("max retries,")
							break
					except ConnectionResetError:
						time.sleep(maxsleep)
						print("Server ConnectionError. Retrying")
				
				for _axie in axies:
					match_genes = GetGenes256(_axie['genes'])
					##check if d == r1
					match_mouth = match_genes.mouth.dId == match_genes.mouth.r1Id
					match_horn = match_genes.horn.dId == match_genes.horn.r1Id
					match_back = match_genes.back.dId == match_genes.back.r1Id
					match_tail = match_genes.tail.dId == match_genes.tail.r1Id
					print(match_mouth, "match mouth")
					print(match_horn, "match horn")
					print(match_back, "match back")
					print(match_tail, "match tail")
					print("\n\n\n\n")

					#perfect genes
					match_perfect_mouth = match_genes.mouth.dId == match_genes.mouth.r1Id == match_genes.mouth.r2Id
					match_perfect_horn = match_genes.horn.dId == match_genes.horn.r1Id == match_genes.horn.r2Id
					match_perfect_back = match_genes.back.dId == match_genes.back.r1Id == match_genes.back.r2Id
					match_perfect_tail = match_genes.tail.dId == match_genes.tail.r1Id == match_genes.tail.r2Id
					if _axie['auction']['seller'] == _axie['owner']:
						if d_r1:
							if ((match_perfect_mouth >= main_axie_perfect_mouth) and (match_perfect_horn >= main_axie_perfect_horn) and (match_perfect_back >= main_axie_perfect_back) and (match_perfect_tail >= main_axie_perfect_tail)):
								#if match's d == r1 is true, and main axie's d == r1 is also false; then returns true
								#if main axie's d== r1 and match's d != r1, return false
								#True = 1, False = 0 for reference
								check_axie = list(filter(lambda a: a['id'] == _axie['id'], r1_list))
								if check_axie == []:
									r1_list.append(_axie)
								else:
									print("Axie already in list.")
								#check temp list size
								if len(r1_list) >= 10:
									print("r1_list full")
									break_out = True
									break

							elif ((match_mouth >= main_axie_better_mouth and not main_axie_perfect_mouth) and (match_horn >= main_axie_better_horn and not main_axie_perfect_horn) and (match_back >= main_axie_better_back and not main_axie_perfect_back) and (match_tail >= main_axie_better_tail and not main_axie_perfect_tail)):
								#if match's d == r1 is true, and main axie's d == r1 is also false; then returns true
								#if main axie's d== r1 and match's d != r1, return false
								#True = 1, False = 0 for reference

								check_axie = list(filter(lambda a: a['id'] == _axie['id'], r1_list))
								if check_axie == []:
									r1_list.append(_axie)
								else:
									print("Axie already in list.")
								#check temp list size
								if len(r1_list) >= 10:
									print("r1_list full")
									break_out = True
									break

				if break_out == True:
					break
	
	r1_query.append({"query": query, "data": r1_list})




	if total <= 0:
		pass
	else:
		if r_response != []:
			r = r_response[0]['data']
			print("found in query")
		else:
			for res in market_response['data']['axies']['results']:
				if len(r) >= 10:
					break
				if res['auction']['seller'] == res['owner']:
					r.append(res)
			r_query.append({"query": query, "data": r})
	if d_r1:
		if len(r1_list) <= 0:
			d_r1_average_eth = ""
			d_r1_average_usd = ""
			min_d_r1_eth = ""
		else:
			d_r1_average_eth = sum([convert_to_eth(float(x['auction']['currentPrice'])) for x in r1_list]) / len(r1_list)
			d_r1_average_usd = sum([float(x['auction']['currentPriceUSD']) for x in r1_list]) / len(r1_list)
			min_d_r1_eth = convert_to_eth(r1_list[0]['auction']['currentPrice'])
	else:
		d_r1_average_eth = ""
		d_r1_average_usd = ""
		min_d_r1_eth = ""



	if len(r) <= 0:
		average_eth = ""
		average_usd = ""
		min_eth = ""
	else:
		average_eth = sum([convert_to_eth(float(x['auction']['currentPrice'])) for x in r]) / len(r)
		average_usd = sum([float(x['auction']['currentPriceUSD']) for x in r]) / len(r)
		min_eth = convert_to_eth(r[0]['auction']['currentPrice'])

	max_row = ws.max_row
	print('Row #', max_row)
	print(axie['id'])
	ws.cell(row = max_row + 1, column = 1).value = axie['id']
	ws.cell(row = max_row + 1, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie['id']}/"
	if axie['auction'] is None:
		ws.cell(row = max_row + 1, column = 1).font = hyperlinked
	else:
		ws.cell(row = max_row + 1, column = 1).font = striked
	ws.cell(row = max_row + 1, column = 2).value = axie['owner']
	ws.cell(row = max_row + 1, column = 2).alignment = Alignment(horizontal = "right")
	ws.cell(row = max_row + 1, column = 3).value = min_eth
	ws.cell(row = max_row + 1, column = 3).number_format = "0.000"
	ws.cell(row = max_row + 1, column = 4).value = average_eth
	ws.cell(row = max_row + 1, column = 4).number_format = "0.000"
	ws.cell(row = max_row + 1, column = 5).value = average_usd
	ws.cell(row = max_row + 1, column = 5).number_format = "0.000"
	ws.cell(row = max_row + 1, column = 6).value = total
	ws.cell(row = max_row + 1, column = 7).value = axie['breedCount']
	ws.cell(row = max_row + 1, column = 8).value = min_d_r1_eth
	ws.cell(row = max_row + 1, column = 8).number_format = "0.000"
	ws.cell(row = max_row + 1, column = 9).value = d_r1_average_eth
	ws.cell(row = max_row + 1, column = 9).number_format = "0.000"
	for i, pair in enumerate(r):
		ws.cell(row = max_row + 1, column = 10 + (i * 2)).value = pair['id']
		ws.cell(row = max_row + 1, column = 10 + (i * 2)).hyperlink = f"https://marketplace.axieinfinity.com/axie/{pair['id']}/"
		ws.cell(row = max_row + 1, column = 10 + (i * 2)).font = hyperlinked
		ws.cell(row = max_row + 1, column = 11 + (i * 2)).value = convert_to_eth(pair['auction']['currentPrice'])
		ws.cell(row = max_row + 1, column = 11 + (i * 2)).number_format = "0.000"
	ws.cell(row = max_row + 1, column = 30).value = axie['class']
	ws.cell(row = max_row + 1, column = 31).value = genes.horn.dName.title()
	ws.cell(row = max_row + 1, column = 32).value = genes.mouth.dName.title()
	ws.cell(row = max_row + 1, column = 33).value = genes.back.dName.title()
	ws.cell(row = max_row + 1, column = 34).value = genes.tail.dName.title()
	ws.cell(row = max_row + 1, column = 35).value = datetime.datetime.utcnow()
	ws.cell(row = max_row + 1, column = 36).value = "UTC"
	ws.cell(row = max_row + 1, column = 37).value = "Click Here"
	part_string = [f"&part={x}" for x in parts]
	part_string = "".join(part_string)
	class_string = f"?class={axie['class']}"
	breed_string = "&breedCount=0&breedCount=0"
	hp_string = "".join([f"&hp={x}" for x in hp])
	morale_string = "".join([f"&morale={x}" for x in morale])
	skill_string = "".join([f"&skill={x}" for x in skill])
	speed_string = "".join([f"&speed={x}" for x in speed])
	ws.cell(row = max_row + 1, column = 37).hyperlink = f"https://marketplace.axieinfinity.com/axie/{class_string}{part_string}{breed_string}{speed_string}{hp_string}{morale_string}{skill_string}"

tday = datetime.datetime.today().strftime("%m_%d_%Y")
dirlist = os.listdir("outputs")
pile = [x for x in dirlist if f"axie_marketplace_{tday}" in x]
pile = [x.replace(f"axie_marketplace_{tday}_", "").replace(".xlsx", "").replace("~$", "") for x in pile]
counts = [int(x) for x in pile]
if len(counts) <= 0:
	d = 1
else:
	d = max(counts) + 1

ws.auto_filter.ref = ws.dimensions
for i, row in enumerate(ws.iter_rows(min_row = 2)):
	if i % 2 == 0:
		for cell in row:
			cell.fill = PatternFill(fill_type="solid", start_color = "c8f4fa", end_color = "c8f4fa")
wb.save(f"outputs/axie_marketplace_{tday}_{d}.xlsx")

if past_5:
	with open("json_files/market_json.json", "w") as outfile:
		json.dump(r_query, outfile)

	with open("json_files/market_json_d_r1.json", "w") as outfile:
		json.dump(r1_query, outfile)



print(f"Runtime {datetime.datetime.now() - start_time}")
