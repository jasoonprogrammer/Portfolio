from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, Http404
from .forms import *
from .models import *
import json
from django.forms import formset_factory, modelformset_factory
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.decorators import login_required, user_passes_test
import datetime
from django.db.models import Max, Sum
from django.contrib import messages
from django.forms.models import model_to_dict
from pyzbar.pyzbar import decode
from PIL import Image
import qrcode
import os
from django.views.generic.list import ListView
from django.views.generic.edit import UpdateView
from django.contrib.auth.mixins import UserPassesTestMixin, LoginRequiredMixin
from django.views import View




import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import math
import threading
from openpyxl.styles import NamedStyle

def is_ajax(request):
	return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

# wb = openpyxl.load_workbook("C:/users/user/desktop/acc_codes.xlsx")
# ws = wb['ACCOUNT CODES']
# for i in range(3, ws.max_row + 1):
# 	title = ws.cell(row = i, column = 1).value
# 	code = ws.cell(row = i, column = 2).value
# 	if code is not None:
# 		AccountCode.objects.create(title = title.strip(), code = code)


def is_doc_fac(user):
	try:
		return user.userlevel.level.name == "Document Facilitator"
	except:
		return user.is_superuser

def is_inv_in_charge(user):
	return user.userlevel.level.name == "Inventory-in-Charge" or user.is_superuser

def is_ins_officer(user):
	return user.userlevel.level.name == "Inspection Officer" or user.is_superuser

def is_superuser(user):
	return user.is_superuser


def change_options(request):
	if is_ajax(request):
		po_id = request.GET['po_id']
		po = PurchaseOrder.objects.get(pk = po_id)
		inspections = po.inspection_set.all()
		items = [inspection.item_set.all() for inspection in inspections]
		item_ids = []
		for item in items:
			for x in item:
				item_ids.append(x.id)
		ims = InventoryMonitoring.objects.filter(item__pk__in = item_ids)
		users = {im.remarks for im in ims}
		context = {}
		context['po'] = po
		context['users'] = users
		return render(request, "newapp/po_list_option.html", context)

@login_required
def po_monitoring(request):
	if is_ajax(request):
		po_no = request.GET['po_no']
		po = [x for x in PurchaseOrder.objects.all() if x.po_no == po_no]
		if len(po) != 0:
			dic = model_to_dict(po[0])
			s = Supplier.objects.get(pk = dic['supplier'])
			dic['supplier_name'], dic['supplier_contact'] = s.name, s.contact
			dic['requesting_end_user_name'] = RequestingEndUser.objects.get(pk = dic['requesting_end_user']).name

			return JsonResponse(dic, safe = False)
		else:
			return JsonResponse({'status': 'fail'}, safe = False)
	if request.user.is_superuser:
		pass
	elif request.user.userlevel.level.name != "Document Facilitator":
		raise Http404("Your account is not allowed to access this page.")
	if request.method == "POST":
		form = PurchaseOrderForm(request.POST)
		if form.is_valid():
			x = form.save()
			messages.success(request, f"You have successfully encoded the Purchase Order Info for {x}")
			return redirect("home")
	else:
		form = PurchaseOrderForm()
	return render(request, "newapp/purchase_order_monitoring.html", {'form': form})


class HomeView(LoginRequiredMixin, View):

	def get(self, request, *args, **kwargs):
		AccountCodeFormset = modelformset_factory(AccountCode, fields = ("code", "title"), extra = 0)
		formset = AccountCodeFormset(queryset = AccountCode.objects.filter(title = ""))
		account_titles = AccountCode.objects.filter(title = "")
		print(account_titles.count())
		context = {}
		context['formset'] = formset
		context['account_titles'] = account_titles
		return render(request, 'newapp/home.html', context)

	def post(self, request, *args, **kwargs):
		AccountCodeFormset = modelformset_factory(AccountCode, fields = ("code", "title"), extra = 0)
		formset = AccountCodeFormset(request.POST, queryset = AccountCode.objects.filter(title = ""))
		if formset.is_valid():
			for form in formset:
				form = form.save()
		context = {}
		context['formset'] = formset
		return redirect("home")


@login_required
@user_passes_test(is_superuser)
def new(request):
	if request.method == "POST":
		form = UserCreationForm(request.POST)
		level_form = LevelForm(request.POST)
		if form.is_valid() and level_form.is_valid():
			f = form.save()
			l = level_form.save(commit = 0)
			l.user = f
			l.save()
			messages.success(request, f'You have successfully registered an account for {f.username}')
			return redirect("new")
	user_form = UserCreationForm()
	level_form = LevelForm()
	return render(request, 'newapp/new.html', {"user_form": user_form, "level_form": level_form})

def get_contact(request):
	if is_ajax(request):
		s = Supplier.objects.filter(name = request.GET['supplier_name'])
		if s.exists():
			s = s.first()
			return HttpResponse(f"{s.contact}")
		return HttpResponse("")

def ins_po(request):
	if is_ajax(request):
		po_no = request.GET['po_no']
		po = po_no_to_po(po_no)
		if po is not None:
			po = model_to_dict(po)
			po['supplier_name'] = Supplier.objects.get(pk = po['supplier']).name
			return JsonResponse(po, safe = False)
		return JsonResponse({"None": None}, safe = False)

@login_required
def inspection_form(request):
	if request.user.is_superuser:
		pass
	elif request.user.userlevel.level.name != "Inspection Officer":
		raise Http404("Your account is not allowed to access this page.")
	if request.method == "POST":
		form = InspectionForm(request.POST)
		if form.is_valid():
			messages.success(request, f'You have successfully encoded Inspection Form')
			item_nos = request.POST.getlist('item_no_list')
			descs = request.POST.getlist('item_desc_list')
			item_qtys = request.POST.getlist('item_qty_list')
			unit_ids = request.POST.getlist('item_unit_id_list')
			item_costs = request.POST.getlist('item_cost_list')
			receipt_nos = request.POST.getlist("receipt_no_list")
			receipt_types = request.POST.getlist("receipt_type_list")
			type_ids = request.POST.getlist('type_id_list')
			receipt_dates = request.POST.getlist("receipt_date_list")
			f = form.save()
			all_ins = Inspection.objects.filter(date__year = datetime.date.today().year, campus__id = request.POST['campus'])
			max_series = all_ins.aggregate(Max('series'))['series__max']
			if max_series is None:
				max_series = 1
			else:
				max_series += 1
			f.series = max_series
			f.save()
			for i, element in enumerate(item_nos):
				item = Item(no = item_nos[i], desc = descs[i], unit = Unit.objects.get(pk = unit_ids[i]), cost = item_costs[i], inspection = f, qty = item_qtys[i])
				item.save()
			for i, element in enumerate(receipt_nos):
				receipt = Receipt(no = receipt_nos[i], type = ReceiptType.objects.get(pk = type_ids[i]), date = receipt_dates[i], inspection = f)
				receipt.save()

			return redirect("inspection_form")
		else:
			context = {}
			context['form'] = form
			context['table1'] = render(request, "newapp/invoice_and_dr.html")
			context['table2'] = render(request, "newapp/items.html")
			context['receipt_type'] = ReceiptType.objects.all()
			context['units'] = Unit.objects.all()
			return render(request, "newapp/inspection_form.html", context)
	else:
		form = InspectionForm()
		context = {}
		context['form'] = form
		context['receipt_type'] = ReceiptType.objects.all()
		context['units'] = Unit.objects.all()
		return render(request, "newapp/inspection_form.html", context)
@login_required
def none(request):
	if is_ajax(request):
		return render(request, "newapp/invoice_and_dr.html")
@login_required
def none2(request):
	if is_ajax(request):
		return render(request, "newapp/items.html")
@login_required
def receipt_row(request):
	if is_ajax(request):
		receipt_no = request.GET['receipt_no']
		receipt_type = ReceiptType.objects.get(pk = request.GET['receipt_type'])
		receipt_date = request.GET['receipt_date']
		context = {}
		context['receipt_no'] = receipt_no
		context['receipt_type'] = receipt_type
		context['receipt_date'] = receipt_date

		return render(request, "newapp/receipt_row.html", context)
@login_required
def generate_inspection_no(request):
	if is_ajax(request):
		campus = Campus.objects.get(pk = request.GET['campus'])
		all_ins = Inspection.objects.filter(date__year = datetime.datetime.today().year, campus = campus).all().aggregate(Max("series"))
		max_series = all_ins['series__max']
		if max_series is None:
			max_series = 1
		else:
			max_series += 1
		stringified = f'{datetime.datetime.now().strftime("%m%d%y")}-{str(max_series).zfill(3)}({campus.name[0].title()}I)'
		return HttpResponse(stringified)
@login_required
def item_row(request):
	if is_ajax(request):
		context = {}
		unit = Unit.objects.get(pk = request.GET['unit'])
		item_no = request.GET['item_no']
		desc = request.GET['desc']
		qty = request.GET['qty']
		cost = request.GET['cost']
		context['unit'] = unit
		context['item_no'] = item_no
		context['desc'] = desc
		context['qty'] = qty
		context['cost'] = cost
		context['total'] = float(cost) * int(qty)
		context['total'] = "{:.2f}".format(context['total'])
		return render(request, 'newapp/item_row.html', context)
@login_required
def set_account_codes_list(request):
	ins_list = Inspection.objects.all()
	context = {}
	context['ins_list'] = ins_list
	return render(request, "newapp/set_acc_code_list.html", context)

@login_required
def set_account_codes_ins(request, pk):
	if request.method == "POST":
		items = request.POST.getlist('item_id')
		articles = request.POST.getlist('article')
		acc_codes = request.POST.getlist('acc_code_id')
		input_acc_codes = request.POST.getlist("acc_code")
		input_acc_title = request.POST.getlist("acc_title")
		article_ids = request.POST.getlist("article_id")
		classifications = request.POST.getlist("classifications")
		for i, item in enumerate(request.POST.getlist('item_id')):
			c_item = Item.objects.get(pk = item)
			article = Article.objects.filter(name = articles[i])
			if not article.exists():
				article = Article.objects.create(name = articles[i])
			else:
				article = article.first()
			c_item.article = article
			print(input_acc_codes[i])
			if acc_codes[i] != "":
				acc_code = AccountCode.objects.get(pk = acc_codes[i])
			else:
				if not AccountCode.objects.filter(code = acc_codes[i]).exists():
					acc_code = AccountCode.objects.create(code = input_acc_codes[i], title = input_acc_title[i])
				else:
					acc_code = AccountCode.objects.filter(pk = acc_codes[i]).first()
			c_item.classification = Classification.objects.get(pk = classifications[i])
			c_item.acc_code = acc_code
			c_item.save()
			ins = c_item.inspection
			ins.with_acc_codes = True
			ins.save()
		messages.success(request, "You have successfully added Account Codes and Articles to the Items.")
		return redirect('acc_code_list')
	acc_codes = AccountCode.objects.all().order_by('title')
	context = {}
	par_class = Classification.objects.filter(name = "PAR").first()
	ics_class = Classification.objects.filter(name = "ICS").first()
	ris_class = Classification.objects.filter(name = "RIS").first()

	items = Inspection.objects.get(pk = pk).item_set.all().order_by("no")
	articles = Article.objects.all().order_by("name")
	context['items'] = items
	context['acc_codes'] = acc_codes
	context['articles'] = articles
	context['par_class'] = par_class
	context['ics_class'] = ics_class
	context['ris_class'] = ris_class
	return render(request, 'newapp/set_acc_code_ins.html', context)

@login_required
def get_acc_info(request):
	pk = request.GET['acc_id']
	acc = AccountCode.objects.get(pk = pk)
	data = {'code': acc.code, 'title': acc.title, 'acc_code_id': pk}
	return JsonResponse(data, safe = False)

@login_required
def get_acc_info2(request):
	acc_code = request.GET['acc_code']
	if AccountCode.objects.filter(code = acc_code).exists():
		acc = AccountCode.objects.filter(code = acc_code).first()
		print("aa")
		data = {'code': acc.code, 'title': acc.title, 'acc_code_id': acc.pk}
		return JsonResponse(data, safe = False)
	else:
		data = {'code': False}
		return JsonResponse(data, safe = False)


@login_required
def inventory_monitoring_form(request):
	if request.method == "POST":

		item = request.POST.getlist("item")
		# # series = request.POST.getlist("series")
		# acc_codes = request.POST.getlist("acc_code")
		# print(acc_codes)
		balance = request.POST.getlist("balance")
		onhand  = request.POST.getlist("onhand")
		remarks = request.POST.getlist('remarks')
		shortage  = request.POST.getlist("shortage")
		designation  = request.POST.getlist("designation")
		date = request.POST.getlist("date")
		wmr = request.POST.getlist("wmr")
		purpose = request.POST.getlist("purpose")
		new_user = request.POST.getlist("new_user")
		unit_count = request.POST.getlist("unit_count")
		date_transferred = request.POST.getlist("date_transferred")
		equipment_status = request.POST.getlist("equipment_status")
		campus_ids = request.POST.getlist("campus")
		for i, element in enumerate(item):
			chosen_item = Item.objects.get(pk = item[i])
			acode = chosen_item.acc_code
			for x in range(int(balance[i])):
				im = InventoryMonitoring()
				im.item = Item.objects.get(pk = item[i])
				im.series = im.max_series(acode)
				im.balance = 1
				im.onhand = 1
				im.designation = designation[i]
				r = ItemUser.objects.filter(name = remarks[i])
				if r.exists():
					im.remarks = r.first()
				else:
					im.remarks = ItemUser.objects.create(name = remarks[i])
				if new_user[i] != "":
					new_user_name = new_user[i]
					neu = NewEndUser(name = new_user_name)
					if unit_count[i] != "":
						neu.count = unit_count[i]
					if date_transferred[i] != "":
						neu = date_transferred[i]
					neu.save()
					im.transferred_to_new_user = neu
				im.wmr = wmr[i]
				
				im.user_campus = Campus.objects.get(pk = campus_ids[i])
				im.save()
				if im.item.classification.name == "PAR":
					p_num = ParNumber.objects.filter(po = im.item.inspection.po, item_user = im.remarks)
					if p_num.exists():
						im.par_num = p_num.first()
						im.save()
					else:
						po_year = im.item.inspection.po.date.year
						last_num = ParNumber.objects.aggregate(Max('series'))['series__max']
						if last_num is None:
							last_num = 1
						else:
							last_num += 1
						p = ParNumber.objects.create(po = im.item.inspection.po, item_user = im.remarks, series = last_num)
						im.par_num = p
						im.save()
			p_no = im.property_no
			qr = qrcode.make(p_no)
			qr.save(f"newapp/static/qrs/{p_no}.png")
		messages.success(request, f"You have successfully encoded {len(item)} item/s.")
		return redirect("inventory-list")
		# balance
		# onhand
		# shortage
		# remarks
		# series
		# wmr
		# equipment_status
		# transferred_to_new_user
		# date_transferred
		# purpose
		# designation
		# date
	return render(request, "newapp/inventory_monitoring_form.html")
@login_required
def monitoring_row_form(request):
	po_no = request.GET['po_no']
	po = po_no_to_po(po_no)
	if po is None:
		return HttpResponse(status = 400)
	items = []
	dates = []
	for ins in po.inspection_set.all():
		dates.append(ins.date_received)
		for item in ins.item_set.all():
			items.append(item)

	context = {}
	context['items'] = items
	context['default_campus'] = po.inspection_set.all().first().campus
	context['dates'] = dates
	context['po'] = po
	context['no'] = request.GET['no']
	context['campuses'] = Campus.objects.all()
	return render(request, 'newapp/monitoring_form_add.html', context)

def po_no_to_po(po_no):
	try:
		md = po_no.split("-")[0]
		y = po_no.split("-")[2]
		mdy = f'{md}{y}'
		d = datetime.datetime.strptime(mdy, "%m%d%y")
		series = po_no.split("-")[1]
		po = PurchaseOrder.objects.filter(date = d, series = series)
		if po.exists():
			return po.first()
		else:
			return None
	except IndexError:
		return None
	except ValueError:
		return None


@login_required
def get_po_info(request):
	po_no = request.GET['po_no']
	po = po_no_to_po(po_no)
	items = []
	dates = []
	for ins in po.inspection_set.all():
		dates.append(ins.date_received)
		for item in ins.item_set.all():
			items.append(item)
	context = {}
	context['items'] = items
	context['dates'] = dates
	return render(request, 'newapp/monitoring_form_add.html', context)
@login_required
def get_property_tag(request):
	if is_ajax(request):
		item_id = request.GET['id']
		item = Item.objects.get(pk = item_id)
		data = model_to_dict(item)
		data['supplier'] = item.inspection.po.supplier.name
		item = Item.objects.filter(pk = request.GET['id'])
		if not item.exists():
			return HttpResponse(status = 400)
		item = item.first()
		article = item.article
		invs = InventoryMonitoring.objects.filter(item__acc_code = item.acc_code, item__inspection__date_received__year = item.inspection.date_received.year)
		c_series = request.GET['current_series']
		if c_series == "":
			max_series = invs.aggregate(Max('series'))['series__max']
			if max_series is None:
				data['acc_series'] = 1
			else:
				max_series += 1
				data['acc_series'] = max_series
		else:
			data['acc_series'] = int(c_series) + 1
		data['acc_id'] = item.acc_code.id
		data['acc_code'] = item.acc_code.code
		data['date'] = item.inspection.date_received.strftime('%m%y')
		data['school'] = item.inspection.campus.name[0]
		data['article'] = item.article.name
		data['unit'] = Unit.objects.get(pk = data['unit']).name
		return JsonResponse(data, safe = False)

@login_required
def po_list(request):
	context = {}
	r = [x for x in range(1, 1001)]
	if request.method == "POST":
		form = POFilter(request.POST)
		if form.is_valid():
			context['results'] = form.search()
			context['filtered'] = True
	else:
		form = POFilter()
	context['range'] = r
	context['form'] = form
	context['cyear'] = datetime.datetime.today().year
	return render(request, "newapp/po_list.html", context)

@login_required
def inventory_list(request):
	inventory = InventoryMonitoring.objects.filter(onhand__gt = 0, balance__gt = 0).order_by("-id")
	context = {}
	context['objs'] = inventory
	context['form'] = InventoryFilter()
	if request.method == "POST":
		form = InventoryFilter(request.POST)
		if form.is_valid():
			context['objs'] = form.search()
		context['form'] = form
		if is_ajax(request):
			if request.POST.getlist("checked") == [""]:
				context['objs'] = InventoryMonitoring.objects.all().order_by("-date")
			else:
				checked = request.POST.getlist("checked")[0].split(",")
				context['checked'] = InventoryMonitoring.objects.filter(id__in = checked)
				context['objs'] = context['objs'].exclude(id__in = checked)
				context['objs'] = context['objs'].filter(balance__gt = 0, onhand__gt = 0).order_by("-date")
			return render(request, "newapp/re_inventory_list.html", context)
	return render(request, "newapp/inventory_list.html", context)

def sticker(request, pk):
	obj = get_object_or_404(InventoryMonitoring, pk = pk)
	url = obj.property_no + ".png"
	context = {}
	context['obj'] = obj
	context['url'] = url
	return render(request, "newapp/sticker.html", context)

def scan(request):
	context = {}
	context['form'] = Scanner()
	if request.method == "POST":
		form = Scanner(request.POST, request.FILES)
		if form.is_valid():
			x = form.save()
			img = Image.open(x.img.url[1:])
			try:
				d = decode(img)
				p_no = d[0].data.decode("ascii")
				obj = [x for x in InventoryMonitoring.objects.all() if p_no == x.property_no]
				obj  = obj[0]
				ScanModel.objects.get(pk = x.id).delete()
				return redirect("info", pk = obj.id)
			except IndexError:
				raise Http404("Item not Found")
	return render(request, "newapp/scan.html", context)

def info(request, pk):
	obj = get_object_or_404(InventoryMonitoring, pk = pk)
	context = {}
	context['obj'] = obj
	return render(request, "newapp/info.html", context)

date_style = NamedStyle(name='datetime', number_format='DD/MM/YYYY')

def print_par(request):
	par = ParNumber.objects.get(pk = request.GET['par_id'])
	rows = []
	for i in range(9, 50):
		rows.append(i)
	for i in range(72, 113):
		rows.append(i)
	for i in range(135, 176):
		rows.append(i)
	for i in range(198, 239):
		rows.append(i)
	for i in range(261, 302):
		rows.append(i)

	print(par)
	par_no = par.par_no
	items = InventoryMonitoring.objects.filter(par_num = par, item__classification__name = "PAR")
	print(items)
	eu = par.item_user
	pairs = []
	for item in items:
		words = word_breaker(item.item.desc + " " + item.property_no, 32)

		for i, word in enumerate(words, start = 1):
			if i == 1:
				pairs.append((item.item.no, item.balance, item.item.unit.name, word, item.item.inspection.date_received, item.item.cost))
			else:
				pairs.append(("", "", "", word, "", ""))
	extra = 1 + ((len(pairs) - 44) / 43)
	pages = math.ceil(extra)
	print(pages)
	wb = openpyxl.load_workbook("newapp/static/PAR.xlsx")
	ws = wb['Sheet1']


	for i in range(0, 6):
		e_r = 63 * i
		# ws.cell(row = 52 + e_r, column = 3).value =f"PO No."
		# ws.cell(row = 57 + e_r , column = 1).value = eu.name
		# ws.cell(row = 59 + e_r, column = 1).value = item.designation
		# ws.cell(row = 61 + e_r, column = 1).value = item.date
		# ws.cell(row = 61 + e_r, column = 1).style = date_style

		ws.cell(row = 7 + e_r, column = 8).value = "PAR#:"
		ws.cell(row = 7 + e_r, column = 8).alignment = Alignment(horizontal = "right")
		ws.cell(row = 7 + e_r, column = 9).value = par_no
		ws.cell(row = 7 + e_r, column = 9).alignment = Alignment(horizontal = "center")

	for i, pair in enumerate(pairs):
		ws.cell(row = rows[i], column = 1).value = pair[0]
		ws.cell(row = rows[i], column = 2).value = pair[1]
		ws.cell(row = rows[i], column = 1).alignment = Alignment(horizontal = "center")
		ws.cell(row = rows[i], column = 2).alignment = Alignment(horizontal = "center")
		ws.cell(row = rows[i], column = 1).font = Font(name = "Segoe UI Light", size = 10)
		ws.cell(row = rows[i], column = 2).font = Font(name = "Segoe UI Light", size = 10)
		ws.cell(row = rows[i], column = 3).value = pair[2]
		ws.cell(row = rows[i], column = 3).alignment = Alignment(horizontal = "center")
		ws.cell(row = rows[i], column = 3).font = Font(name = "Segoe UI Light", size = 10)
		ws.merge_cells(start_row = rows[i], end_row = rows[i], start_column = 4, end_column = 6)
		ws.cell(row = rows[i], column = 4).value = pair[3]
		ws.cell(row = rows[i], column = 4).font = Font(name = "Anonymous Pro", size = 9)
		ws.cell(row = rows[i], column = 4).number_format = "mm/dd/yy"
		ws.cell(row = rows[i], column = 4).alignment = Alignment(horizontal = "center")
		ws.cell(row = rows[i], column = 7).value = pair[4]
		ws.cell(row = rows[i], column = 7).font = Font(name = "Segoe UI Light", size = 10)
		ws.cell(row = rows[i], column = 7).alignment = Alignment(horizontal = "center")
		ws.cell(row = rows[i], column = 7).number_format = "mm/dd/yy"
		ws.cell(row = rows[i], column = 8).value = pair[5]
		ws.cell(row = rows[i], column = 8).font = Font(name = "Segoe UI Light", size = 10)
		ws.cell(row = rows[i], column = 8).alignment = Alignment(horizontal = "center")
		ws.cell(row = rows[i], column = 8).number_format = '#,##0.00'


	# ws.print_area = f"A1:I{63 * pages}"
	rrr = 63 * pages
	ws.print_area = f"A1:I{rrr}"
	wb.save(f"newapp/static/PAR_{par_no}.xlsx")
	os.system(f"newapp\\static\\PAR_{par_no}.xlsx")
	return render(request, "newapp/printing.html")

def print_ins(request, ins_no):
	rows = []
	for i in range(12, 57):
		rows.append(i)
	for i in range(82, 126):
		rows.append(i)
	for i in range(151, 195):
		rows.append(i)
	for i in range(220, 264):
		rows.append(i)
	for i in range(289, 333):
		rows.append(i)
	for i in range(358, 402):
		rows.append(i)
	for i in range(427, 471):
		rows.append(i)
	for i in range(496, 540):
		rows.append(i)
	for i in range(565, 609):
		rows.append(i)
	for i in range(634, 678):
		rows.append(i)

	ins = Inspection.objects.get(pk = ins_no)
	items = ins.item_set.all()
	pairs = []
	for item in items:
		words = word_breaker(item.desc, 46)
		for i, word in enumerate(words, start = 1):
			if i == 1:
				pairs.append((item.no, item.unit.name, word, item.qty, item.cost))
			else:
				pairs.append(("", "", word, "", ""))
	extra = len(pairs) / 44
	pages = math.ceil(extra)
	wb = openpyxl.load_workbook("newapp/static/inspection_form.xlsx")
	ws = wb['Sheet1']
	for i in range(0, 11):
		e_r = 69 * i
		ws.cell(row = 5 + e_r, column = 1).value = ins.ins_no
		ws.cell(row = 8 + e_r, column = 3).value = ins.po.supplier.name
		ws.cell(row = 9 + e_r, column = 3).value = ins.po.po_no
		ws.cell(row = 9 + e_r, column = 4).value = ins.po.date
		invoice = ins.receipt_set.filter(type__name = "Invoice")
		if invoice.exists():
			invoice = invoice.first()
			ws.cell(row = 9 + e_r, column = 6).value = invoice.no
			ws.cell(row = 9 + e_r, column = 8).value = invoice.date
		ws.cell(row = 10 + e_r, column = 4).value = ins.requisitioning_office.name
		ws.cell(row = 59 + e_r, column = 3).value = ins.date
		ws.cell(row = 59 + e_r, column = 7).value = ins.date_received
		ws.cell(row = 67 + e_r, column = 1).value = ins.inspection_officer.name
	for i, pair in enumerate(pairs):
		ws.cell(row = rows[i], column = 1).value = pair[0]
		ws.cell(row = rows[i], column = 2).value = pair[1]
		ws.merge_cells(start_row = rows[i], end_row = rows[i], start_column = 3, end_column = 5)
		ws.cell(row = rows[i], column = 3).value = pair[2]
		ws.cell(row = rows[i], column = 3).font = Font(name = "Anonymous Pro")
		ws.cell(row = rows[i], column = 3).alignment = Alignment(horizontal = "center")
		ws.cell(row = rows[i], column = 6).value = pair[3]
		ws.cell(row = rows[i], column = 7).value = pair[4]

	ws.print_area = f"A1:H{69 * pages}"
	print(ins.id)
	wb.save(f"newapp/static/inspection_id_{ins.id}.xlsx")
	os.system(f"newapp\\static\\inspection_id_{ins.id}.xlsx")
	return render(request, f"newapp/printing.html")

def word_breaker(sentence, break_len):
	sentence_arr = sentence.split()
	w_len = 0
	broken_words = []
	c_words = ''
	for word in sentence_arr:
		if w_len > break_len:
			w_len = 0
			broken_words.append(c_words)
			c_words = ""
			w_len += len(word)
			if c_words != "" and w_len + len(word) < break_len:
				c_words += " " + word
			else:
				c_words += word
		else:
			w_len += len(word) + 1
			if c_words != "":
				c_words += " " + word
			else:
				c_words += word
	broken_words.append(c_words)
	return broken_words

def copy_cells(ws, start_row, start_column, end_row, end_column, x_offset = 0, y_offset = 0):
	# ws.merge_cells(start_row = 5 + x_offset, end_row = 5 + x_offset, start_column = 1, end_column = 2)
	for r in range(start_row, end_row):
		for c in  range(start_column, end_column):
			if r % 52 in [1, 2, 3, 4]:
				if c == 1:
					ws.merge_cells(start_row = r + x_offset, end_row = r + x_offset, start_column = 1, end_column = 8)
					ws.cell(row = r + x_offset, column = c + y_offset).value = ws.cell(row = r, column = c).value
					ws.cell(row = r+ x_offset, column = c + y_offset).alignment = Alignment(horizontal = "center")
					ws.cell(row = r+ x_offset, column = c + y_offset).font = Font(name = "Segoe UI Light", size = 10)
				if r % 52 == 1 or r % 52 == 4:
					ws.row_dimensions[r + x_offset].height = 40.5
					ws.cell(row = r + x_offset, column = c + y_offset).font = Font(size = 12)
					if r % 52 == 4:
						ws.cell(row = r + x_offset, column = c + y_offset).alignment = Alignment(vertical = "top", horizontal = "center")
				if r % 52 == 2:
					ws.cell(row = r + x_offset, column = c + y_offset).font = Font(bold = True)


			else:
				if r % 52 == 5 and c == 1:
					ws.merge_cells(start_row = r + x_offset, end_row = r + x_offset, start_column = 1, end_column = 2)
					ws.cell(row = r + x_offset, column = 1).value = ws.cell(row = r, column = c).value
				else:
					ws.cell(row = r + x_offset, column = c + y_offset).value = ws.cell(row = r, column = c).value

class InspectionListView(ListView, UserPassesTestMixin):

	model = Inspection
	template_name = "newapp/inspection_list.html"
	context_object_name = "objects"

	def test_func(self):
		return user.userlevel.level.name == "Inspection Officer" or user.is_superuser

def update_inspection(request, pk):
	context = {}
	ins = Inspection.objects.get(pk = pk)
	x = model_to_dict(ins)
	x['requisitioning_office'] = ins.requisitioning_office.name
	x['supplier_name'] = ins.po.supplier.name
	x['date_inspected'] = x['date'].strftime("%Y-%m-%y")
	ReceiptFormset = modelformset_factory(Receipt, exclude = ['inspection'], extra = 0)
	ItemFormset = modelformset_factory(Item, exclude = ['inspection'], extra = 0)
	if request.method == "POST":
		form = InspectionUpdateForm(request.POST)
		context['form'] = form
		if form.is_valid():
			form.save()
		messages.success(request, "Inspection Information Updated Successfully")
		return redirect("inspection-list")
	else:
		context['form'] = InspectionUpdateForm(x)

	context['ins'] = ins
	return render(request, "newapp/update_inspection.html", context)

def update_inspection(request, pk):
	context = {}
	ins = Inspection.objects.get(pk = pk)
	x = model_to_dict(ins)
	x['requisitioning_office'] = ins.requisitioning_office.name
	x['supplier_name'] = ins.po.supplier.name
	x['date_inspected'] = x['date'].strftime("%Y-%m-%y")
	ReceiptFormset = modelformset_factory(Receipt, exclude = ['inspection'], extra = 0)
	ItemFormset = modelformset_factory(Item, exclude = ['inspection'], extra = 0)
	if request.method == "POST":
		form = InspectionUpdateForm(request.POST)
		context['form'] = form
		if form.is_valid():
			form.save()
		messages.success(request, "Inspection Information Updated Successfully")
		return redirect("inspection-list")
	else:
		context['form'] = InspectionUpdateForm(x)

	context['ins'] = ins
	return render(request, "newapp/update_inspection.html", context)

def delete_inspection(request, pk):
	inspection = Inspection.objects.get(pk = pk)
	inspection.delete()
	return redirect('inspection-list')

def print_inventory(request):
	checked = request.GET.getlist("checked[]")
	ims = InventoryMonitoring.objects.filter(pk__in = checked).order_by("-date")
	wb = openpyxl.load_workbook("newapp/static/inventory_blank.xlsx")
	ws = wb['Sheet1']
	border = Border(left = Side(border_style="dotted", color = "000000"), right = Side(border_style="dotted", color = "000000"))

	for i, item in enumerate(ims, start = 10):
		ws.cell(row = i, column = 1).value = item.date
		ws.cell(row = i, column = 1).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 1).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 1).number_format = "mm/dd/yy"
		ws.cell(row = i, column = 1).border = border
		ws.cell(row = i, column = 2).value = item.item.article.name
		ws.cell(row = i, column = 2).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 2).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 2).border = border
		ws.cell(row = i, column = 3).value = item.item.desc
		ws.cell(row = i, column = 3).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 3).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 3).border = border
		ws.cell(row = i, column = 4).value = item.property_no
		ws.cell(row = i, column = 4).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 4).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 4).border = border
		ws.cell(row = i, column = 5).value = item.item.unit.name
		ws.cell(row = i, column = 5).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 5).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 5).border = border
		ws.cell(row = i, column = 6).value = item.item.cost
		ws.cell(row = i, column = 6).alignment = Alignment(horizontal = "right")
		ws.cell(row = i, column = 6).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 6).number_format = '#,##0.00'
		ws.cell(row = i, column = 6).border = border
		ws.cell(row = i, column = 7).value = item.balance
		ws.cell(row = i, column = 7).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 7).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 7).border = border
		ws.cell(row = i, column = 8).value = item.onhand
		ws.cell(row = i, column = 8).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 8).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 8).border = border
		ws.cell(row = i, column = 9).value = item.shortage
		ws.cell(row = i, column = 9).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 9).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 9).border = border
		ws.cell(row = i, column = 10).value = item.overage
		ws.cell(row = i, column = 10).alignment = Alignment(horizontal = "right")
		ws.cell(row = i, column = 10).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 10).number_format = '#,##0.00'
		ws.cell(row = i, column = 10).border = border
		ws.cell(row = i, column = 11).value = item.remarks.name
		ws.cell(row = i, column = 11).alignment = Alignment(horizontal = "center")
		ws.cell(row = i, column = 11).font = Font(name = "Segoe UI Light", size = 9)
		ws.cell(row = i, column = 11).border = border
		
	pages = 1 + math.ceil(ims.count() / 25)

	wb.save("newapp/static/inventory_1.xlsx")
	os.system(f"newapp\\static\\inventory_1.xlsx")
	return render(request, "newapp/printing.html")





def print_ics(request):
	po_no = request.GET['po_id']
	rows = []
	for i in range(9, 57):
		rows.append(i)
	for i in range(79, 126):
		rows.append(i)
	for i in range(148, 195):
		rows.append(i)
	for i in range(217, 264):
		rows.append(i)

	ins = Inspection.objects.filter(po__pk = po_no)
	items = list()
	for x in ins:
		for item in x.item_set.all():
			items.append(item)

	pairs = []
	for item in items:
		words = word_breaker(item.desc, 51)
		for i, word in enumerate(words, start = 1):
			if i == 1:
				pairs.append((item.no, item.qty, item.unit.name, word, item.inspection.date, item.cost))
			else:
				pairs.append(("", "", word, "", ""))
	extra = len(pairs) / 44
	pages = math.ceil(extra)
	wb = openpyxl.load_workbook("newapp/static/ICS.xlsx")
	ws = wb['Sheet1']

	# for i in range(0, 11):
	# 	e_r = 69 * i
	# 	ws.cell(row = 5 + e_r, column = 1).value = ins.ins_no
	# 	ws.cell(row = 8 + e_r, column = 3).value = ins.po.supplier.name
	# 	ws.cell(row = 9 + e_r, column = 3).value = ins.po.po_no
	# 	ws.cell(row = 9 + e_r, column = 4).value = ins.po.date
	# 	invoice = ins.receipt_set.filter(type__name = "Invoice")
	# 	if invoice.exists():
	# 		invoice = invoice.first()
	# 		ws.cell(row = 9 + e_r, column = 6).value = invoice.no
	# 		ws.cell(row = 9 + e_r, column = 8).value = invoice.date
	# 	ws.cell(row = 10 + e_r, column = 4).value = ins.requisitioning_office.name
	# 	ws.cell(row = 59 + e_r, column = 3).value = ins.date
	# 	ws.cell(row = 59 + e_r, column = 7).value = ins.date_received
	# 	ws.cell(row = 67 + e_r, column = 1).value = ins.inspection_officer.name


	for i, pair in enumerate(pairs):
		ws.cell(row = rows[i], column = 1).value = pair[0]
		ws.cell(row = rows[i], column = 2).value = pair[1]
		ws.cell(row = rows[i], column = 3).value = pair[2]
		ws.cell(row = rows[i], column = 4).value = pair[2]
		ws.cell(row = rows[i], column = 7).value = pair[3]
		ws.cell(row = rows[i], column = 8).value = pair[4]

	ws.print_area = f"A1:H{69 * pages}"
	wb.save(f"newapp/static/ICS_{ins.po.po_no}.xlsx")
	os.system(f"newapp\\static\\ICS_{ins.po.po_no}.xlsx")
	return render(request, f"newapp/printing.html")

def transfer_items(request):
	if request.method == "POST":
		ids = request.POST.getlist("obj_id")
		qtys = request.POST.getlist("obj_qty")
		to_designation = request.POST['to_designation']
		print(to_designation)
		statuses = request.POST.getlist("obj_status")
		qtys = [int(x) for x in qtys]
		to_id = request.POST.get("hidden_to")
		if to_id == "":
			item_user = ItemUser(name = request.POST.get('transfer_to'))
			item_user.save()
		else:
			item_user = ItemUser.objects.get(pk = to_id)
		par_no = ParNumber(item_user = item_user)
		par_no.series = par_no.max_series(datetime.datetime.today().year)
		par_no.save()
		zipped = zip(ids, qtys)
		for i, v in enumerate(zipped):
			obj = InventoryMonitoring.objects.get(pk = v[0])
			obj.onhand -= v[1]
			obj.balance -= v[1]
			obj2 = InventoryMonitoring.objects.get(pk = v[0])
			obj2.pk = None
			obj2.save()
			obj2.onhand = v[1]
			obj2.balance = v[1]
			obj2.from_par = obj.par_num
			obj2.par_num = par_no
			obj2.designation = to_designation
			obj2.equipment_status = statuses[i]
			obj2.date = datetime.datetime.today()
			obj2.remarks = item_user
			obj2.save()
			obj.save()
		return redirect("transfer-items")
	if is_ajax(request):
		from_person = request.GET['from_id']
		person = ItemUser.objects.get(pk = from_person)
		context = {}
		items = InventoryMonitoring.objects.filter(remarks = person, balance__gte = 1)
		context['items'] = items
		return render(request, "newapp/transfer_items_rows.html", context)
	context = {}
	context['persons'] = ItemUser.objects.all()
	return render(request, "newapp/transfer_items.html", context)

def confirm_transfer(request):
	context = {}
	context['objects'] = []
	arr = json.loads(request.GET['json_arr'])
	for key, val in arr.items():
		obj = InventoryMonitoring.objects.get(pk = key)
		md = model_to_dict(obj)
		md['qty'] = val[0]
		md['property_no'] = obj.property_no
		md['desc'] = obj.item.desc
		md['status'] = val[1]
		context['objects'].append(md)


	return render(request, "newapp/transfer_confirm.html", context)


#check if its transferred to an existing itemuser
def name_transfer_autocheck(request):
	name = request.GET['name']
	item_user = ItemUser.objects.filter(name = name)
	if item_user.exists():
		item_user = item_user.first()
		item_user = model_to_dict(item_user)
		item_user['existing'] = True
		return JsonResponse(item_user, safe = False)
	else:
		return JsonResponse({'existing': False}, safe = False)

def property_timeline(request, property_no):
	context = {}
	last = InventoryMonitoring.objects.get(pk = property_no)
	p_no = last.property_no
	objs = [last]
	while last.from_par is not None:
		last = InventoryMonitoring.objects.filter(par_num = last.from_par)
		last = [x for x in last if x.property_no == p_no]
		last = last[0]
		objs.append(last)
		pks = [x.id for x in objs]
	context['objs'] = objs
	if len(objs) > 0:
		obj = objs[0]
		context['obj'] = obj
	return render(request, 'newapp/timeline.html', context)

def delete_po(request):
	if is_ajax(request):
		po = PurchaseOrder.objects.get(id = request.GET['po_id'])
		po.delete()
		return JsonResponse({}, safe = False)

def par_list(request):
	context = {}
	pars = ParNumber.objects.all()
	context['pars'] = pars
	context['accounts'] = AccountCode.objects.all()
	return render(request, "newapp/par_list.html", context)


def accounting(request):
	if request.method == "POST":
		codes = request.POST.getlist("codes[]")
		wb = openpyxl.load_workbook("newapp/static/_accounting_template.xlsx")
		ws = wb['Sheet1']
		ws.title = "Accounting"
		left_right_border = Border(left = Side(border_style="thin", color="000000"), right=Side(border_style="thin", color="000000"))
		bottom_border = Border(bottom  = Side(border_style="thin", color="000000"),
			right  = Side(border_style="thin", color="000000"),
			left  = Side(border_style="thin", color="000000"))
		from_date = datetime.datetime.strptime(request.POST['start_date'], "%Y-%m-%d")
		end_date = datetime.datetime.strptime(request.POST['end_date'], "%Y-%m-%d")
		date_today = end_date.strftime("%B %d, %Y").upper()

		string = f"As of {date_today}"
		side_no = Side(border_style = None)
		side_yes = Side(border_style = "thin", color="000000")
		no_border = Border(left= side_no, right= side_no, bottom= side_no, top= side_no)
		all_border = Border(left = side_yes, right = side_yes, bottom = side_yes, top = side_yes)
		ws.cell(row = 2, column = 1).value = string
		for i, code in enumerate(codes, start = 8):
			ws.cell(row = i, column = 1).value = AccountCode.objects.get(pk = code).code
			ws.cell(row = i, column = 1).border = left_right_border
			ws.cell(row = i, column = 2).value = AccountCode.objects.get(pk = code).title
			ws.cell(row = i, column = 2).border = left_right_border
			ims = InventoryMonitoring.objects.filter(item__acc_code__id = code, user_campus__name = "Talisay", date__gte = from_date, date__lte = end_date)
			arr_ims = [x.total_val for x in ims]
			vals = []
			ws.cell(row = i, column = 3).value = sum(arr_ims)
			vals.append(sum(arr_ims))
			ws.cell(row = i, column = 3).number_format = '#,##0.00'
			ims = InventoryMonitoring.objects.filter(item__acc_code__id = code, user_campus__name = "Fortune Towne", date__gte = from_date, date__lte = end_date)
			arr_ims = [x.total_val for x in ims]
			ws.cell(row = i, column = 4).value = sum(arr_ims)
			vals.append(sum(arr_ims))
			ws.cell(row = i, column = 4).number_format = '#,##0.00'
			ims = InventoryMonitoring.objects.filter(item__acc_code__id = code, user_campus__name = "Alijis", date__gte = from_date, date__lte = end_date)
			arr_ims = [x.total_val for x in ims]
			ws.cell(row = i, column = 5).value = sum(arr_ims)
			vals.append(sum(arr_ims))
			ws.cell(row = i, column = 5).number_format = '#,##0.00'
			ims = InventoryMonitoring.objects.filter(item__acc_code__id = code, user_campus__name = "Binalbagan", date__gte = from_date, date__lte = end_date)
			arr_ims = [x.total_val for x in ims]
			ws.cell(row = i, column = 6).value = sum(arr_ims)
			vals.append(sum(arr_ims))
			ws.cell(row = i, column = 6).number_format = '#,##0.00'
			ims = InventoryMonitoring.objects.filter(item__acc_code__id = code, user_campus__name = "Fermin Research & Extension", date__gte = from_date, date__lte = end_date)
			arr_ims = [x.total_val for x in ims]
			ws.cell(row = i, column = 7).value = sum(arr_ims)
			vals.append(sum(arr_ims))
			ws.cell(row = i, column = 7).number_format = '#,##0.00'
			ws.cell(row = i, column = 8).value = sum(vals)
			ws.cell(row = i, column = 8).number_format = '#,##0.00'
			if i % 25 == 7:
				ws.cell(row = i, column = 1).border = bottom_border
				ws.cell(row = i, column = 2).border = bottom_border
				ws.cell(row = i, column = 3).border = bottom_border
				ws.cell(row = i, column = 4).border = bottom_border
				ws.cell(row = i, column = 5).border = bottom_border
				ws.cell(row = i, column = 6).border = bottom_border
				ws.cell(row = i, column = 7).border = bottom_border
				ws.cell(row = i, column = 8).border = bottom_border
			else:
				ws.cell(row = i, column = 1).border = left_right_border
				ws.cell(row = i, column = 2).border = left_right_border
				ws.cell(row = i, column = 3).border = left_right_border
				ws.cell(row = i, column = 4).border = left_right_border
				ws.cell(row = i, column = 5).border = left_right_border
				ws.cell(row = i, column = 6).border = left_right_border
				ws.cell(row = i, column = 7).border = left_right_border
				ws.cell(row = i, column = 8).border = left_right_border
		last_row = len(codes) + 8
		ws.cell(row = last_row, column = 1).value = "Total"
		ims = InventoryMonitoring.objects.filter(item__acc_code__id__in = codes, user_campus__name = "Talisay", date__gte = from_date, date__lte = end_date)
		arr_ims = [x.total_val for x in ims]
		ws.cell(row = last_row, column = 3).value = sum(arr_ims)
		ws.cell(row = last_row, column = 3).number_format = '#,##0.00'
		ims = InventoryMonitoring.objects.filter(item__acc_code__id__in = codes, user_campus__name = "Fortune Towne", date__gte = from_date, date__lte = end_date)
		arr_ims = [x.total_val for x in ims]
		ws.cell(row = last_row, column = 4).value = sum(arr_ims)
		ws.cell(row = last_row, column = 4).number_format = '#,##0.00'
		ims = InventoryMonitoring.objects.filter(item__acc_code__id__in = codes, user_campus__name = "Alijis", date__gte = from_date, date__lte = end_date)
		arr_ims = [x.total_val for x in ims]
		ws.cell(row = last_row, column = 5).value = sum(arr_ims)
		ws.cell(row = last_row, column = 5).number_format = '#,##0.00'
		ims = InventoryMonitoring.objects.filter(item__acc_code__id__in = codes, user_campus__name = "Binalbagan", date__gte = from_date, date__lte = end_date)
		arr_ims = [x.total_val for x in ims]
		ws.cell(row = last_row, column = 6).value = sum(arr_ims)
		ws.cell(row = last_row, column = 6).number_format = '#,##0.00'
		ims = InventoryMonitoring.objects.filter(item__acc_code__id__in = codes, user_campus__name = "Fermin Research & Extension", date__gte = from_date, date__lte = end_date)
		arr_ims = [x.total_val for x in ims]
		ws.cell(row = last_row, column = 7).value = sum(arr_ims)
		ws.cell(row = last_row, column = 7).number_format = '#,##0.00'
		ims = InventoryMonitoring.objects.filter(item__acc_code__id__in = codes, date__gte = from_date, date__lte = end_date)
		arr_ims = [x.total_val for x in ims]
		ws.cell(row = last_row, column = 8).value = sum(arr_ims)
		ws.cell(row = last_row, column = 8).number_format = '#,##0.00'
		ws.cell(row = last_row, column = 1).border = all_border
		ws.cell(row = last_row, column = 2).border = all_border
		ws.cell(row = last_row, column = 3).border = all_border
		ws.cell(row = last_row, column = 4).border = all_border
		ws.cell(row = last_row, column = 5).border = all_border
		ws.cell(row = last_row, column = 6).border = all_border
		ws.cell(row = last_row, column = 7).border = all_border
		ws.cell(row = last_row, column = 8).border = all_border


		wb.save(f"newapp/static/accounting_1.xlsx")
		os.system("start EXCEL.EXE newapp\\static\\accounting_1.xlsx")

	context = {}
	accounts = AccountCode.objects.all()
	context['accounts'] = accounts
	return render(request, "newapp/accounting.html", context)

class POEncodingView(View):

	def get(self, request):
		context = {}
		encoding_po = POEncodingForm()
		context['encoding_po'] = encoding_po
		return render(request, "newapp/encoding_po.html", context)

	def post(self, request):
		context = {}
		encoding_po = POEncodingForm(request.POST)
		encoding_inspection = InspectionEncodingForm()
		if encoding_po.is_valid():
			po = encoding_po.save()
			return redirect("po_encoding")
		context['encoding_po'] = encoding_po
		return render(request, "newapp/encoding_po.html", context)

class InspectionEncodingView(View):

	def get(self, request):
		context = {}
		context['title'] = "Inspection Encoding"
		encoding_inspection = InspectionEncodingForm()
		context['encoding_inspection'] = encoding_inspection
		context['receipt_form'] = ReceiptEncodingForm()
		return render(request, "newapp/encoding_inspection.html", context)

	def post(self, request):
		context = {}
		context['title'] = "Inspection Encoding"
		encoding_inspection = InspectionEncodingForm(request.POST)
		if encoding_inspection.is_valid():
			encoding_inspection = encoding_inspection.save()
			receipt_form = ReceiptEncodingForm(request.POST)
			# receipt_form.inspection = encoding_inspection
			# if receipt_form.is_valid():
			# 	receipt_form.save()
			# context['receipt_form'] = receipt_form
			# context['encoding_inspection'] = encoding_inspection
			# # return render(request, "newapp/encoding_inspection.html", context)

			return redirect("inspection_encoding")


class ItemEncodingFormsetView(View):

	def get(self, request):
		context = {}
		item_count = request.GET['count']

		ItemFormSet = formset_factory(ItemEncodingForm, extra = int(item_count))

		item_formset = ItemFormSet()
		context['formset'] = item_formset
		context['item_count'] = [x for x in range(int(item_count))]
		context['units'] = Unit.objects.all()
		context['classifications'] = Classification.objects.all()
		return render(request, "newapp/encoding_item_formset.html", context)

class ItemEncodingView(View):

	def get(self, request):
		context = {}
		return render(request, "newapp/encoding_item.html", context)

	def post(self, request):
		context = {}
		for i in range(len(request.POST.getlist("item-no"))):
			x = Item()
			x.no = request.POST.getlist("item-no")[i]
			all_article = [z.name for z in Article.objects.all()]
			article = request.POST.getlist("article")[i]
			if article in all_article:
				x.article = Article.objects.filter(name = article).first()
			else:
				x.article = Article.objects.create(name = article)
			x.desc = request.POST.getlist("desc")[i]
			x.unit = Unit.objects.get(pk = request.POST.getlist("unit")[i])
			x.qty = request.POST.getlist("qty")[i]
			x.cost = request.POST.getlist("cost")[i]
			all_inspection = [z.ins_no for z in Inspection.objects.all()]
			inspection = request.POST.getlist("inspection")[i]
			if inspection in all_inspection:
				x.inspection = [z for z in Inspection.objects.all() if z.ins_no == inspection]
			all_acc_code = [z.code for z in AccountCode.objects.all()]
			acc_code = request.POST.getlist("account-code")[i]
			if acc_code in all_acc_code:
				acc_code = AccountCode.objects.filter(code = request.POST.getlist("account-code")[i]).first()
			else:
				acc_code = AccountCode.objects.create(code = acc_code, title="")

			x.acc_code = acc_code
			x.classification = Classification.objects.get(pk = request.POST.getlist("classification")[i])
			x.save()
			return redirect("item_encoding")


		return render(request, "newapp/encoding_item.html", context)


