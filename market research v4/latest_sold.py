import requests, urllib, json, os, datetime
from dependencies.market_search_modules import start_soldquery
import time
from dependencies.mymodules import convert_to_eth
from openpyxl.styles import PatternFill, Border, Side, Font
import openpyxl
from dependencies.gene_class import GetGenes256

def main_function():
	start_time = time.time()
	max_retries = 3
	sleep_time = 1.5
	retries = 0
	wb = openpyxl.load_workbook("outputs/latest_sold.xlsx")
	ws = wb['Latest']
	ws2 = wb['Sale History']
	side_border = Border(left = Side(style = "thin"), right = Side(style = "thin"))
	hyperlinked = Font(underline = "single", color="0C10F5")

	json_file = open("json_files/latest_sold_axies.json")
	json_response = json.load(json_file)
	json_file.close()
	result_no = 1

	def profile_link(ronin):
		ronin_account = ronin.replace("0x", "ronin:")
		link = f"https://marketplace.axieinfinity.com/profile/{ronin_account}/axie/"
		return link

	current_index = 0
	break_out = False
	#loop until error disappears
	while True:
		try:
			while True:
				print(f"Loop # {current_index}")
				#size is for query size
				data = start_soldquery(size = 10000, index = 20 * current_index)
				time.sleep(sleep_time)
				response_data = data['data']
				settled_auctions = response_data['settledAuctions']['axies']
				total = settled_auctions['total']
				results = settled_auctions['results']
				# print("number of results", len(results))




				for i, axie in enumerate(results):
					gene_error = False
					try:
						genes = GetGenes256(axie['genes'])
					except KeyError:
						gene_error = True
					if i % 2 == 0:
						fill = PatternFill(fill_type="solid", start_color = "c8f4fa", end_color = "c8f4fa")
					else:
						fill = PatternFill(fill_type="solid", start_color = "ffffff", end_color = "ffffff")

					initial_row = 2
					initial_row2 = 2
					axie_id = axie['id']
					transfer_history = axie['transferHistory']
					total_transfers = transfer_history['total']
					sale_history = transfer_history['results']
					#get the latest result to get the owner and seller
					latest_history = sale_history[0]
					_from = latest_history['from']
					_to = latest_history['to']
					timestamp = latest_history['timestamp']
					tx_hash = axie['transferHistory']['results'][0]['txHash']
					result_no += 1


					#if it is already in json list, break the loop
					if tx_hash in json_response:
						print(f"{tx_hash} found in list, stopping the loop")
					else:

						print(tx_hash)
						print("Axie Result #", result_no)
						json_response.append(tx_hash)

						ws.insert_rows(2)
						ws2.insert_rows(2)

						#latest sold sheet
						eth_price = convert_to_eth(latest_history['withPrice'])
						usd_price = latest_history['withPriceUsd']
						ws.cell(row = initial_row, column = 1).value = axie_id
						ws.cell(row = initial_row, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie_id}/"
						ws.cell(row = initial_row, column = 1).font = hyperlinked
						ws.cell(row = initial_row, column = 1).fill = fill
						ws.cell(row = initial_row, column = 2).value = _from
						ws.cell(row = initial_row, column = 2).hyperlink = profile_link(_from)
						ws.cell(row = initial_row, column = 2).font = hyperlinked
						ws.cell(row = initial_row, column = 2).fill = fill
						ws.cell(row = initial_row, column = 3).value = _to
						ws.cell(row = initial_row, column = 3).hyperlink = profile_link(_to)
						ws.cell(row = initial_row, column = 3).font = hyperlinked
						ws.cell(row = initial_row, column = 3).fill = fill
						ws.cell(row = initial_row, column = 4).value = eth_price
						ws.cell(row = initial_row, column = 4).number_format = "0.000"
						ws.cell(row = initial_row, column = 4).fill = fill
						ws.cell(row = initial_row, column = 5).value = float(usd_price)
						ws.cell(row = initial_row, column = 5).number_format = "0.000"
						ws.cell(row = initial_row, column = 5).fill = fill
						if not gene_error:
							ws.cell(row = initial_row, column = 6).value = axie['class']
							ws.cell(row = initial_row, column = 7).value = genes.horn.dName.title()
							ws.cell(row = initial_row, column = 8).value = genes.mouth.dName.title()
							ws.cell(row = initial_row, column = 9).value = genes.back.dName.title()
							ws.cell(row = initial_row, column = 10).value = genes.tail.dName.title()
							ws.cell(row = initial_row, column = 11).value = genes.purity
							ws.cell(row = initial_row, column = 12).value = tx_hash

						ws.cell(row = initial_row, column = 6).fill = fill
						ws.cell(row = initial_row, column = 7).fill = fill
						ws.cell(row = initial_row, column = 8).fill = fill
						ws.cell(row = initial_row, column = 9).fill = fill
						ws.cell(row = initial_row, column = 10).fill = fill
						ws.cell(row = initial_row, column = 11).fill = fill


						#history sheet
						ws2.cell(row = initial_row2, column = 1).value = axie_id
						ws2.cell(row = initial_row2, column = 1).hyperlink = f"https://marketplace.axieinfinity.com/axie/{axie_id}/"
						ws2.cell(row = initial_row2, column = 1).font = hyperlinked
						ws2.cell(row = initial_row2, column = 1).fill = fill
						ws2.cell(row = initial_row2, column = 2).value = _from
						ws2.cell(row = initial_row2, column = 2).hyperlink = profile_link(_from)
						ws2.cell(row = initial_row2, column = 2).font = hyperlinked
						ws2.cell(row = initial_row2, column = 2).fill = fill
						ws2.cell(row = initial_row2, column = 3).value = _to
						ws2.cell(row = initial_row2, column = 3).hyperlink = profile_link(_to)
						ws2.cell(row = initial_row2, column = 3).font = hyperlinked
						ws2.cell(row = initial_row2, column = 3).fill = fill
						ws2.cell(row = initial_row2, column = 4).value = eth_price
						ws2.cell(row = initial_row2, column = 4).number_format = "0.000"
						ws2.cell(row = initial_row2, column = 4).fill = fill
						ws2.cell(row = initial_row2, column = 5).value = float(usd_price)
						ws2.cell(row = initial_row2, column = 5).number_format = "0.000"
						ws2.cell(row = initial_row2, column = 5).fill = fill
						if not gene_error:
							ws2.cell(row = initial_row2, column = 7).value = genes.horn.dName.title()
							ws2.cell(row = initial_row2, column = 8).value = genes.mouth.dName.title()
							ws2.cell(row = initial_row2, column = 9).value = genes.back.dName.title()
							ws2.cell(row = initial_row2, column = 10).value = genes.tail.dName.title()
							ws2.cell(row = initial_row2, column = 11).value = genes.purity
							ws2.cell(row = initial_row2, column = 6).value = axie['class']
							
						ws2.cell(row = initial_row2, column = 6).fill = fill
						ws2.cell(row = initial_row2, column = 7).fill = fill
						ws2.cell(row = initial_row2, column = 8).fill = fill
						ws2.cell(row = initial_row2, column = 9).fill = fill
						ws2.cell(row = initial_row2, column = 10).fill = fill
						ws2.cell(row = initial_row2, column = 11).fill = fill

						for index, transfer in enumerate(sale_history[1:], start = 1):
							ws2.insert_rows(initial_row2 + index)
							ws2.cell(row = initial_row2 + index, column = 1).fill = fill
							ws2.cell(row = initial_row2 + index, column = 2).value = transfer['from']
							ws2.cell(row = initial_row2 + index, column = 2).fill = fill
							ws2.cell(row = initial_row2 + index, column = 2).hyperlink = profile_link(transfer['from'])
							ws2.cell(row = initial_row2 + index, column = 2).font = hyperlinked
							ws2.cell(row = initial_row2 + index, column = 3).value = transfer['to']
							ws2.cell(row = initial_row2 + index, column = 3).fill = fill
							ws2.cell(row = initial_row2 + index, column = 3).hyperlink = profile_link(transfer['to'])
							ws2.cell(row = initial_row2 + index, column = 3).font = hyperlinked
							ws2.cell(row = initial_row2 + index, column = 4).fill = fill
							ws2.cell(row = initial_row2 + index, column = 5).fill = fill

							ws2.cell(row = initial_row2 + index, column = 4).value = convert_to_eth(transfer['withPrice'])
							ws2.cell(row = initial_row2 + index, column = 4).number_format = "0.000"

							ws2.cell(row = initial_row2 + index, column = 5).value = float(transfer['withPriceUsd'])
							ws2.cell(row = initial_row2 + index, column = 5).number_format = "0.000"
							ws2.cell(row = initial_row2 + index, column = 6).fill = fill
							ws2.cell(row = initial_row2 + index, column = 7).fill = fill
							ws2.cell(row = initial_row2 + index, column = 8).fill = fill
							ws2.cell(row = initial_row2 + index, column = 9).fill = fill
							ws2.cell(row = initial_row2 + index, column = 10).fill = fill
							ws2.cell(row = initial_row2 + index, column = 11).fill = fill



				current_index += 1
				if break_out:
					break


				#check if it is the first run, if it is, then run only the first 100 axies
				if current_index >= 5 or len(results) == 0:
					break_out = True
					break
				break
			if break_out:
				break

		# except TypeError:
		# 	retries += 1
		# 	print(f"Something went wrong. Retrying: {retries} times")
		# 	time.sleep(sleep_time)
		# 	if retries <= max_retries:
		# 		print("Max retries. Please rerun")
		# 		break
		except ConnectionResetError:
			retries += 1
			print(f"Server Error. Retrying: {retries} times")
			time.sleep(sleep_time)
			if retries <= max_retries:
				print("Max retries. Please rerun")
		except ConnectionError:
			retries += 1
			print(f"Server Error. Retrying: {retries} times")
			time.sleep(sleep_time)
			if retries <= max_retries:
				print("Max retries. Please rerun")

	for row in ws.iter_rows(min_row = 2):
		for cell in row:
			cell.border = side_border

	# get all axies from temporary list and add to json_response to add to json file


	ws.auto_filter.ref = ws.dimensions
	ws2.auto_filter.ref = ws2.dimensions

	wb.save(f"outputs/latest_sold.xlsx")


	with open("json_files/latest_sold_axies.json", "w") as outfile:
		json.dump(json_response, outfile)
	print(f"Runtime: {time.time() - start_time}")

if __name__ == "__main__":
	main_function()
	input("Finished. Press enter to close.")
